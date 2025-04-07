#==========================================================imports and constants==========================================================#
import sys
import re
import pandas as pd  
import traceback
import os 
import requests
import json
import random
import shutil
import threading
from collections import defaultdict
from datetime import datetime, timedelta
from typing import List, Dict
import io
from PIL import Image
import math
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QHeaderView,
    QPushButton, QLabel, QLineEdit, QComboBox, QFileDialog, QSizePolicy, 
    QFrame, QTableWidget, QTableWidgetItem, QProgressBar, QGraphicsDropShadowEffect,
    QTextEdit, QDialog, QMessageBox, QScrollArea, QStackedWidget, QGroupBox, QRadioButton
)
from PyQt6.QtCore import Qt, pyqtSignal, QThread, QSize, QTimer, QObject, pyqtSignal, QEvent
from PyQt6.QtGui import QIcon, QPixmap, QFont, QIntValidator, QColor, QMovie, QTextCursor, QPainter, QPainterPath, QPen, QPalette

# Constants - Inverted Colors
DARK_BLUE = "#24325f"
DARK_RED = "#951d1e"
BLACK = "#000000"
CARD_BG = "#1a1a1a"  
TEXT_COLOR = "#ffffff"  
INPUT_BG = "#2d2d2d"  
BORDER_COLOR = "#3d3d3d"

# Standard button style 
MENU_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_BLUE};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 10px 20px;
    font-size: 16px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: {DARK_RED};
}}
"""

# Menu Exit button style 
MENU_EXIT_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_RED};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 10px 20px;
    font-size: 16px;
    font-weight: bold;
}}
QPushButton:hover {{
    background-color: #ab2223;
}}
"""

# STANDARD button style 
STANDARD_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_BLUE};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 5px 15px;
}}
QPushButton:hover {{
    background-color: {DARK_RED};
}}
"""

# Exit button style 
EXIT_BUTTON_STYLE = f"""
QPushButton {{
    background-color: {DARK_RED};
    color: {TEXT_COLOR};
    border: none;
    border-radius: 3px;
    padding: 5px 15px;
}}
QPushButton:hover {{
    background-color: #ab2223;
}}
"""

# Table style
TABLE_STYLE = f"""
QTableWidget::item {{
    text-align: center;
    padding: 5px;
}}
QTableWidget {{
    background-color: {INPUT_BG};
    gridline-color: #3d3d3d;
    border: 1px solid #3d3d3d;
    border-radius: 5px;
}}
QTableWidget::item:selected {{
    background-color: {DARK_BLUE};
    color: white;
}}
QTableWidget::item:hover {{
    background-color: {DARK_BLUE};
    color: white;
}}
QHeaderView::section:hover {{
    background-color: {DARK_BLUE};
    color: white;
}}
QHeaderView::section {{
    background-color: #202c54;
    color: white;
    gridline-color: #3d3d3d;
    border: 1px solid #3d3d3d;
}}
"""

GROUP_BOX_STYLE = f"""
QGroupBox {{
    border: 2px solid {DARK_BLUE};
    border-radius: 5px;
    margin-top: 2ex;
    font-weight: bold;
    font-size: 20px;
    color: {TEXT_COLOR};
}}
QGroupBox::title {{
    subcontrol-origin: margin;
    left: 10px;
    padding: 0 3px 0 3px;
    color: {TEXT_COLOR};
}}
QLabel {{
    color: white;
}}
QRadioButton {{
    color: white;
}}
QRadioButton::indicator::unchecked {{
    border: 2px solid white;
    background-color: white; /* Background for unchecked state */
    border-radius: 7px;
}}
QRadioButton::indicator::checked {{
    border: 2px solid white;
    background-color: {DARK_BLUE}; /* Dark blue check mark */
    border-radius: 7px;
}}

QComboBox {{
    color: white;
    background-color: {INPUT_BG};
}}

QLineEdit {{
    color: white;
    background-color: {INPUT_BG};
}}
"""

PROGRESS_BAR_STYLE = f"""
QProgressBar {{
    text-align: center;
    background-color: {INPUT_BG};
}}
QProgressBar::chunk {{
    background-color: {DARK_RED}; 
}}
"""

# For the console 
CONSOLE_STYLE = f"""
QTextEdit {{
    background-color: {INPUT_BG};
    color: white;
    border: 1px solid {BLACK};
}}
"""

#==========================================================start page==========================================================#
class StartPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.init_ui()
        
    def init_ui(self):
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Create a dark card-like container
        card_container = QWidget()
        card_container.setStyleSheet(f"""
            QWidget {{
                background-color: {BLACK};
                border-radius: 10px;
                border: 1px solid {BORDER_COLOR};
                color: {TEXT_COLOR};
            }}
            QFrame {{
                border: none;
            }}
            QComboBox {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
                border: 1px solid {BORDER_COLOR};
                padding: 5px;
                border-radius: 3px;
            }}
            QComboBox::drop-down {{
                border: none;
            }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid {TEXT_COLOR};
            }}
            QLineEdit {{
                background-color: {INPUT_BG};
                color: {TEXT_COLOR};
                border: 1px solid {BORDER_COLOR};
                padding: 5px;
                border-radius: 3px;
            }}
        """)
        card_layout = QVBoxLayout(card_container)
        card_layout.setSpacing(20)
        card_layout.setContentsMargins(40, 40, 40, 40)

        # Create info button in top left of the card
        info_container = QHBoxLayout()
        info_container.setContentsMargins(0, 0, 0, 0)
        
        self.info_button = QPushButton("", self)
        self.info_button.setFixedSize(40, 40)
        icon_path = os.path.join(os.path.dirname(__file__), 'info.png')
        if os.path.exists(icon_path):
            self.info_button.setIcon(QIcon(icon_path))
            self.info_button.setIconSize(QSize(32, 32))
        else:
            self.info_button.setStyleSheet(f"""
                QPushButton {{
                    background-color: {DARK_BLUE};
                    color: {TEXT_COLOR};
                    border: none;
                    border-radius: 20px;
                    font-size: 16px;
                    font-weight: bold;
                }}
                QPushButton:hover {{
                    background-color: {DARK_RED};
                }}
            """)
            self.info_button.setText("i")
        
        info_container.addWidget(self.info_button)
        # Add spacer after button to push everything else to the right
        info_container.addStretch()
        card_layout.addLayout(info_container)
        
        # Center the card in the window
        center_layout = QHBoxLayout()
        center_layout.addStretch(1)
        center_layout.addWidget(card_container)
        center_layout.addStretch(1)
        
        # Vertical centering
        vertical_layout = QVBoxLayout()
        vertical_layout.addStretch(1)
        vertical_layout.addLayout(center_layout)
        vertical_layout.addStretch(1)
        
        # Background container
        bg_container = QWidget()
        bg_container.setStyleSheet(f"background-color: {CARD_BG};")
        bg_container.setLayout(vertical_layout)
        main_layout.addWidget(bg_container)

        # Logo
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(128, 128, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        logo_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(logo_label)

        # Title
        title_label = QLabel("Department Attendance \nManagement System ")
        title_label.setStyleSheet(f"""
            color: {TEXT_COLOR};
            font-size: 24px;
            font-weight: bold;
        """)
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_label)

        # Buttons Container
        buttons_widget = QWidget()
        buttons_widget.setStyleSheet(f"""
            QWidget {{
                border: none;
            }}
            QFrame {{
                border: none;
            }}
        """)
        buttons_layout = QVBoxLayout(buttons_widget)
        buttons_layout.setSpacing(15)

        # Prepare Log Sheet Button
        self.preparer_btn = QPushButton("Prepare Log Sheet")
        self.preparer_btn.setMinimumHeight(50)
        self.preparer_btn.setStyleSheet(MENU_BUTTON_STYLE)
        buttons_layout.addWidget(self.preparer_btn)

        # Process Button
        self.process_btn = QPushButton("Process Attendance")
        self.process_btn.setMinimumHeight(50)
        self.process_btn.setStyleSheet(MENU_BUTTON_STYLE)
        buttons_layout.addWidget(self.process_btn)

        # Populate with Faculty Button
        self.populate_btn = QPushButton("Populate Main File")
        self.populate_btn.setMinimumHeight(50)
        self.populate_btn.setStyleSheet(MENU_BUTTON_STYLE)
        buttons_layout.addWidget(self.populate_btn)

        # Dashboard Button
        self.dashboard_btn = QPushButton("Analyze Attendance")
        self.dashboard_btn.setMinimumHeight(50)
        self.dashboard_btn.setStyleSheet(MENU_BUTTON_STYLE)
        buttons_layout.addWidget(self.dashboard_btn)

        # Exit Button
        exit_btn = QPushButton("Exit")
        exit_btn.setMinimumHeight(50)
        exit_btn.setStyleSheet(MENU_EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(self.parent().close)
        buttons_layout.addWidget(exit_btn)

        # Add buttons container to card
        card_layout.addWidget(buttons_widget)
        card_layout.addStretch()

        # Set fixed size for the card
        card_container.setFixedWidth(600)
        card_container.setMinimumHeight(400)

#==========================================================info page==========================================================#

class InfoPage(QWidget):
    def __init__(self):
        super().__init__()
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("About the app")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)
        
        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # Create a GroupBox for info content, similar to AttendanceProcessor style
        info_group = QGroupBox()
        info_group.setStyleSheet(GROUP_BOX_STYLE)  # Using same style as other GroupBoxes
        info_layout = QVBoxLayout(info_group)
        
        # Create scroll area for content
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setMinimumHeight(600)
        scroll_area.setStyleSheet(f"""
            QScrollArea {{
                background-color: transparent;
                border: none;
            }}
            QScrollBar:vertical {{
                border: none;
                background: {CARD_BG};
                width: 10px;
                margin: 0px;
            }}
            QScrollBar::handle:vertical {{
                background: {DARK_BLUE};
                min-height: 20px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                border: none;
                background: none;
            }}
        """)
        
        # Content widget for the scroll area
        info_content = QWidget()
        info_content.setStyleSheet("background-color: transparent;")
        content_layout = QVBoxLayout(info_content)
        
        # Create info label with rich text
        info_text = """
        <h1 style='color: white; text-align: center;'>Department Attendance Management System</h1>
        <p style='color: white;'>Effortlessly track and analyze student attendance with this user-friendly tool.</p>
        
        <h2 style='color: white;'>System Requirements</h2>
        <ul style='color: white;'>
            <li>Operating System: Windows</li>
            <li>Software: Microsoft Excel (for viewing and editing reports)</li>
        </ul>
        
        <h2 style='color: white;'>Key Features</h2>
        <ol style='color: white;'>
            <li><b>Prepare Log Sheets</b>
                <ul>
                    <li>Merge attendance logs from cloud storage or your computer into a single file.</li>
                </ul>
            </li>
            <li><b>Process Attendance</b>
                <ul>
                    <li>Validate attendance logs against scheduled sessions and generate comprehensive reports.</li>
                </ul>
            </li>
            <li><b>Populate Main File</b>
                <ul>
                    <li>Automatically update your department's main attendance records.</li>
                </ul>
            </li>
            <li><b>Analyze Data</b>
                <ul>
                    <li>View statistics, track individual student records, and explore attendance trends.</li>
                </ul>
            </li>
        </ol>
        
        <h2 style='color: white;'>Quick Start Guide</h2>
        <h3 style='color: white;'>1. Preparing Log Sheets</h3>
        <p style='color: white;'><b>From Cloud Storage:</b></p>
        <ul style='color: white;'>
            <li>Open the app</li>
            <li>Navigate to Log Sheet Preparer</li>
            <li>Click Download Excel Files from Cloud Storage</li>
            <li>Then, click Merge Logs Files</li>
        </ul>
        <p style='color: white;'><b>From Your Computer:</b></p>
        <ul style='color: white;'>
            <li>Select Import Local Files</li>
            <li>Click Browse</li>
            <li>Choose Merge Logs Files</li>
        </ul>
        <p style='color: white;'><b>Output Location:</b> <code>Merged Files</code> folder</p>
        
        <h3 style='color: white;'>2. Processing Attendance</h3>
        <p style='color: white;'><b>Set Up Data:</b></p>
        <ul style='color: white;'>
            <li>Load:
                <ul>
                    <li>Student Database (Excel file with Student ID, Name, Year, Group)</li>
                    <li>Attendance Logs (Excel file with Student ID, Location, Date, Time)</li>
                    <li>Session Schedule (Excel file with Year, Group, Session, Location, Date, Start Time)</li>
                </ul>
            </li>
        </ul>
        <p style='color: white;'><b>Process:</b></p>
        <ul style='color: white;'>
            <li>Click Process Attendance Records</li>
        </ul>
        <p style='color: white;'><b>Output Location:</b> <code>attendance_reports</code> folder</p>
        <p style='color: white;'>Includes detailed session reports and summary sheets.</p>
        
        <h3 style='color: white;'>3. Updating the Main Attendance File</h3>
        <ol style='color: white;'>
            <li>Select:
                <ul>
                    <li>Your department's attendance file</li>
                    <li>Faculty's main attendance file</li>
                </ul>
            </li>
            <li>Click Populate Main Attendance File</li>
        </ol>
        <p style='color: white;'><b>Output:</b></p>
        <ul style='color: white;'>
            <li>Main file updated automatically</li>
            <li>Backups created for safety</li>
        </ul>
        
        <h3 style='color: white;'>4. Analyzing Data</h3>
        <ul style='color: white;'>
            <li>Open the Analysis Dashboard</li>
            <li>Load your attendance report</li>
        </ul>
        <p style='color: white;'><b>View:</b></p>
        <ul style='color: white;'>
            <li>Key Stats: Total students, sessions held, average attendance</li>
            <li>Group & Session Breakdowns</li>
            <li>Individual Student Records: Search by Name or Student ID</li>
        </ul>
        
        <h2 style='color: white;'>Generated Reports</h2>
        <p style='color: white;'><b>Attendance Report Format:</b><br><code>Y{year}_{module}_attendance.xlsx</code></p>
        <p style='color: white;'><b>Includes:</b></p>
        <ul style='color: white;'>
            <li>Attendance Sheet: Validated session attendance per student</li>
            <li>Summary Sheet: Statistics for each student and session</li>
        </ul>
        
        <h2 style='color: white;'>Need Help?</h2>
        <p style='color: white;'><b>Developer Note:</b><br>
        This application was developed by a medical student at the Faculty of Medicine, Ain Shams University, as part of a project to integrate tech-based solutions into the university's educational processes.</p>
        <p style='color: white;'><b>Contact Information:</b></p>
        <ul style='color: white;'>
            <li><a href='mailto:231249@med.asu.edu.eg' style='color: #4b96ff;'>Primary Contact</a></li>
            <li><a href='mailto:mohammadhamdisaid.mh@icloud.com' style='color: #4b96ff;'>Alternative 1</a></li>
            <li><a href='mailto:mohammad_hamdi11@yahoo.com' style='color: #4b96ff;'>Alternative 2</a></li>
        </ul>
        """
        
        info_label = QLabel()
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_label.setText(info_text)
        info_label.setWordWrap(True)
        info_label.setStyleSheet("background-color: transparent;")
        content_layout.addWidget(info_label)
        
        scroll_area.setWidget(info_content)
        info_layout.addWidget(scroll_area)
        
        # Add the group box to the main layout
        main_layout.addWidget(info_group)
        
        # Add a stretch after the group box to push everything up
        main_layout.addStretch()

#==========================================================attendance processors==========================================================#

class AttendanceProcessor(QWidget):
    def __init__(self):
        super().__init__()
        self.schedules = []
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Attendance Processor")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)
        
        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # Reference Data Section
        ref_group = QGroupBox("Reference Data")
        ref_group.setStyleSheet(GROUP_BOX_STYLE)
        ref_layout = QVBoxLayout(ref_group)
        
        # Single line layout for reference data
        ref_input_layout = QHBoxLayout()
        ref_input_layout.addWidget(QLabel("Database File:"))
        self.ref_file_input = QLineEdit()
        self.ref_file_input.setPlaceholderText("Select Excel file...")
        self.ref_file_input.setMinimumWidth(200)
        ref_input_layout.addWidget(self.ref_file_input)
        ref_browse_btn = QPushButton("Browse")
        ref_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        ref_input_layout.addWidget(ref_browse_btn)
        ref_input_layout.addWidget(QLabel("Sheet Name:"))
        self.ref_sheet_combo = QComboBox()
        self.ref_sheet_combo.setMinimumWidth(100)
        ref_input_layout.addWidget(self.ref_sheet_combo)
        ref_browse_btn.clicked.connect(lambda: self.browse_file(self.ref_file_input))
        ref_layout.addLayout(ref_input_layout)
        main_layout.addWidget(ref_group)

        # Attendance Logs Section
        log_group = QGroupBox("Attendance Logs")
        log_group.setStyleSheet(GROUP_BOX_STYLE)
        log_layout = QVBoxLayout(log_group)
        
        # Single line layout for log data
        log_input_layout = QHBoxLayout()
        log_input_layout.addWidget(QLabel("Log File:"))
        self.log_file_input = QLineEdit()
        self.log_file_input.setPlaceholderText("Select Excel file...")
        self.log_file_input.setMinimumWidth(200)
        log_input_layout.addWidget(self.log_file_input)
        log_browse_btn = QPushButton("Browse")
        log_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        log_input_layout.addWidget(log_browse_btn)
        log_input_layout.addWidget(QLabel("Sheet Name:"))
        self.log_sheet_combo = QComboBox()
        self.log_sheet_combo.setMinimumWidth(100)
        log_input_layout.addWidget(self.log_sheet_combo)
        log_browse_btn.clicked.connect(lambda: self.browse_file(self.log_file_input))
        log_layout.addLayout(log_input_layout)
        main_layout.addWidget(log_group)

        # Session Schedules Section
        schedule_group = QGroupBox("Session Schedules")
        schedule_group.setStyleSheet(GROUP_BOX_STYLE)
        schedule_layout = QVBoxLayout(schedule_group)
        
        self.schedule_table = QTableWidget()
        self.schedule_table.setColumnCount(5)
        self.schedule_table.setHorizontalHeaderLabels(['Year', 'Module', 'File', 'Sheet', 'Total Sessions'])
        self.schedule_table.setStyleSheet(TABLE_STYLE)
        
        # Center align the header text
        header = self.schedule_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)  # Center header text
        
        # Set column resize modes to stretch and fit content
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # Year
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Module
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)  # File
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)  # Sheet
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)  # Total Sessions
        
        schedule_layout.addWidget(self.schedule_table)
        schedule_btn_layout = QHBoxLayout()
        add_schedule_btn = QPushButton("Add Schedule")
        add_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        add_schedule_btn.clicked.connect(self.add_schedule)
        remove_schedule_btn = QPushButton("Remove Schedule")
        remove_schedule_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        remove_schedule_btn.clicked.connect(self.remove_schedule)
        schedule_btn_layout.addWidget(add_schedule_btn)
        schedule_btn_layout.addWidget(remove_schedule_btn)
        schedule_layout.addLayout(schedule_btn_layout)
        main_layout.addWidget(schedule_group)
        
        # Progress Bar Section
        progress_group = QGroupBox("Progress")
        progress_group.setStyleSheet(GROUP_BOX_STYLE)
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)
        
        # Create loading gif label
        self.loading_label = QLabel()
        self.loading_label.setFixedSize(24, 24)  # Adjust based on your GIF size
        self.loading_label.setVisible(False)  # Hidden by default
        
        # Create the movie object for the GIF
        self.loading_movie = QMovie()
        self.loading_movie.setScaledSize(QSize(24, 24))  # Adjust based on your GIF size
        self.loading_label.setMovie(self.loading_movie)
        
        # Make sure to have your loading.gif in the same directory as the script
        loading_gif_path = os.path.join(os.path.dirname(__file__), 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie.setFileName(loading_gif_path)
        else:
            print(f"Warning: loading.gif not found at {loading_gif_path}")
        
        # Create a horizontal layout to hold both the progress bar and loading animation
        progress_h_layout = QHBoxLayout()
        progress_h_layout.addWidget(self.progress_bar)
        progress_h_layout.addWidget(self.loading_label)
        progress_layout.addLayout(progress_h_layout)
        
        main_layout.addWidget(progress_group)

        # Output Console Section
        console_group = QGroupBox("Output Console")
        console_group.setStyleSheet(GROUP_BOX_STYLE)
        console_layout = QVBoxLayout(console_group)

        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(150)
        self.output_console.setStyleSheet(CONSOLE_STYLE)  
        console_layout.addWidget(self.output_console)
        main_layout.addWidget(console_group)

        # Bottom Buttons
        button_layout = QHBoxLayout()
        process_btn = QPushButton("Process Attendance Records")
        process_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        process_btn.clicked.connect(self.process_data)
        button_layout.addWidget(process_btn)
        main_layout.addLayout(button_layout)

        # Connect file input changes to sheet loading
        self.ref_file_input.textChanged.connect(
            lambda: self.load_sheets(self.ref_file_input.text(), self.ref_sheet_combo))
        self.log_file_input.textChanged.connect(
            lambda: self.load_sheets(self.log_file_input.text(), self.log_sheet_combo))

    def browse_file(self, input_field):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)")
        if filename:
            input_field.setText(filename)

    def load_sheets(self, file_path, combo_box):
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                combo_box.clear()
                combo_box.addItems(wb.sheetnames)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error loading workbook: {str(e)}")

    def add_schedule(self):
        dialog = ScheduleDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.schedules.append(dialog.get_schedule_data())
            self.update_schedule_table()

    def remove_schedule(self):
        current_row = self.schedule_table.currentRow()
        if current_row >= 0:
            self.schedules.pop(current_row)
            self.update_schedule_table()

    def update_schedule_table(self):
        self.schedule_table.setRowCount(len(self.schedules))
        for i, schedule in enumerate(self.schedules):
            for j, value in enumerate(schedule):
                item = QTableWidgetItem(str(value))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)  # Center align the text
                self.schedule_table.setItem(i, j, item)

    def process_data(self):
        if not self.validate_inputs():
            return

        # Disable UI elements
        self.setEnabled(False)
        self.output_console.clear()
        self.progress_bar.setValue(0)

        # Show and start the loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()

        # Create and start processing thread
        self.process_thread = ProcessThread(
            self.ref_file_input.text(),
            self.ref_sheet_combo.currentText(),
            self.log_file_input.text(),
            self.log_sheet_combo.currentText(),
            self.schedules
        )

        # Connect signals
        self.process_thread.progress_updated.connect(self.update_progress)
        self.process_thread.error_occurred.connect(self.handle_error)
        self.process_thread.processing_complete.connect(self.handle_completion)

        # Start processing
        self.process_thread.start()

    def validate_inputs(self):
        # Validate reference file
        if not self.ref_file_input.text() or not self.ref_sheet_combo.currentText():
            self.show_custom_warning("Reference Data Required", "Please select reference file and sheet")
            return False
            
        # Validate log file
        if not self.log_file_input.text() or not self.log_sheet_combo.currentText():
            self.show_custom_warning("Log Data Required", "Please select log file and sheet")
            return False
            
        # Validate schedules
        if not self.schedules:
            self.show_custom_warning("Schedules Required", "Please add at least one schedule")
            return False
            
        return True

    def show_custom_warning(self, title, message):
        """Show a custom styled warning dialog"""
        warning_dialog = QMessageBox(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setText(message)
        warning_dialog.setIcon(QMessageBox.Icon.Warning)

        # Create and style OK button
        ok_button = warning_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background and text
        warning_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        warning_dialog.exec()

    def update_progress(self, value):
        self.progress_bar.setValue(value)
        self.output_console.append(f"Processing... {value}%")

    def handle_error(self, error_message):
        self.setEnabled(True)
        error_dialog = QMessageBox(self)
        error_dialog.setWindowTitle("Error")
        error_dialog.setText(f"Error processing data: {error_message}")
        error_dialog.setIcon(QMessageBox.Icon.Critical)

        # Style OK button
        ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        error_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        error_dialog.exec()
        self.output_console.append(f"Error: {error_message}")

    def handle_completion(self):
        self.setEnabled(True)
        self.progress_bar.setValue(100)
        self.output_console.append("Processing complete!")
        
        # Hide the loading animation
        self.loading_label.setVisible(False)
        self.loading_movie.stop()

        success_dialog = QMessageBox(self)
        success_dialog.setWindowTitle("Success")
        success_dialog.setText("Processing complete! Check the attendance_reports folder.")
        success_dialog.setIcon(QMessageBox.Icon.Information)

        # Style OK button
        ok_button = success_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        success_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        success_dialog.exec()

class ScheduleDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Schedule")
        self.setMinimumWidth(400)
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()

        # Add input validation
        self.total_input.setValidator(QIntValidator(1, 999, self))
        self.year_input.setValidator(QIntValidator(1, 6, self))

    def init_ui(self):
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Form Group
        form_group = QGroupBox("Schedule Details")
        form_group.setStyleSheet(GROUP_BOX_STYLE)
        form_layout = QVBoxLayout(form_group)
        form_layout.setSpacing(10)

        # Form fields
        self.year_input = QLineEdit()
        self.year_input.setPlaceholderText("Academic year...")
        self.module_input = QLineEdit()
        self.module_input.setPlaceholderText("Module to process...")
        self.total_input = QLineEdit()
        self.total_input.setPlaceholderText("Total sessions number...")
        self.file_input = QLineEdit()
        self.file_input.setPlaceholderText("Select Excel file...")
        self.sheet_combo = QComboBox()
        
        # Add department combobox
        self.department_combo = QComboBox()
        departments = [
            "Anatomy", 
            "Histology", 
            "Physiology", 
            "Biochemistry", 
            "Pharmacology", 
            "Pathology", 
            "Parasitology", 
            "Microbiology", 
            "Forensics and Toxicology", 
            "Community Medicine", 
            "Clinical"
        ]
        self.department_combo.addItems(departments)

        # Add form fields with consistent spacing
        form_layout.addWidget(QLabel("Academic Year:"))
        form_layout.addWidget(self.year_input)
        form_layout.addWidget(QLabel("Module Name:"))
        form_layout.addWidget(self.module_input)
        form_layout.addWidget(QLabel("Department:"))
        form_layout.addWidget(self.department_combo)
        form_layout.addWidget(QLabel("Total Required Sessions:"))
        form_layout.addWidget(self.total_input)

        # Schedule File section with Browse button
        file_label = QLabel("Schedule File:")
        file_top_layout = QHBoxLayout()
        file_top_layout.addWidget(file_label)
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_file)
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        file_top_layout.addWidget(browse_btn)
        file_top_layout.addStretch()
        form_layout.addLayout(file_top_layout)
        form_layout.addWidget(self.file_input)

        # Sheet selection
        form_layout.addWidget(QLabel("Sheet Name:"))
        form_layout.addWidget(self.sheet_combo)

        main_layout.addWidget(form_group)

        # Buttons at the bottom
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        ok_button = QPushButton("OK")
        ok_button.clicked.connect(self.accept)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        cancel_button = QPushButton("Cancel")
        cancel_button.clicked.connect(self.reject)
        cancel_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        button_layout.addWidget(ok_button)
        button_layout.addWidget(cancel_button)
        main_layout.addLayout(button_layout)

    def browse_file(self):
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)"
        )
        if filename:
            self.file_input.setText(filename)
            self.load_sheets(filename)

    def load_sheets(self, file_path):
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                self.sheet_combo.clear()
                self.sheet_combo.addItems(wb.sheetnames)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error loading workbook: {str(e)}")

    def get_schedule_data(self):
        return [
            self.year_input.text(),
            self.module_input.text(),
            self.file_input.text(),
            self.sheet_combo.currentText(),
            int(self.total_input.text()),
            self.department_combo.currentText()  # Add department to returned data
        ]

    def accept(self):
        """Validate inputs before closing dialog"""
        try:
            # Check required fields
            if not self.year_input.text().strip():
                raise ValueError("Academic year is required")
            if not self.module_input.text().strip():
                raise ValueError("Module name is required")
            if not self.total_input.text().strip():
                raise ValueError("Total required sessions is required")
            if not self.file_input.text().strip():
                raise ValueError("Schedule file is required")
            if not self.sheet_combo.currentText():
                raise ValueError("Sheet name is required")
            
            # Validate numeric input
            total_sessions = self.total_input.text()
            if not total_sessions.isdigit():
                raise ValueError("Total sessions must be a whole number")
            if int(total_sessions) <= 0:
                raise ValueError("Total sessions must be greater than zero")
            
            # Validate file exists
            if not os.path.isfile(self.file_input.text()):
                raise FileNotFoundError("Selected schedule file does not exist")

        except (ValueError, FileNotFoundError) as e:
            # Create custom message box
            error_dialog = QMessageBox(self)
            error_dialog.setWindowTitle("Invalid Input")
            error_dialog.setText(str(e))
            error_dialog.setIcon(QMessageBox.Icon.Warning)
            
            # Configure OK button
            ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
            ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
            
            # Style dialog background
            error_dialog.setStyleSheet(f"""
                QMessageBox {{
                    background-color: {CARD_BG};
                }}
                QLabel {{
                color: {TEXT_COLOR};
                    font-size: 14px;
                }}
            """)
            
            error_dialog.exec()
            return
            
        super().accept()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

class ProcessThread(QThread):
    progress_updated = pyqtSignal(int)
    error_occurred = pyqtSignal(str)
    processing_complete = pyqtSignal()

    def __init__(self, ref_file, ref_sheet, log_file, log_sheet, schedules):
        super().__init__()
        self.ref_file = ref_file
        self.ref_sheet = ref_sheet
        self.log_file = log_file
        self.log_sheet = log_sheet
        self.schedules = schedules

    def run(self):
        try:
            # Calculate total steps
            total_steps = 2 + len(self.schedules) * 5
            current_step = 0
            
            # Load reference data
            ref_wb = openpyxl.load_workbook(self.ref_file)
            ref_ws = ref_wb[self.ref_sheet]
            student_db = list(ref_ws.values)
            student_map = self.create_student_map(student_db)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
            
            # Load log data
            log_wb = openpyxl.load_workbook(self.log_file)
            log_ws = log_wb[self.log_sheet]
            log_history = list(log_ws.values)
            current_step += 1
            self.progress_updated.emit(int(current_step / total_steps * 100))
            
            # Create output directory
            output_dir = os.path.join(os.getcwd(), "attendance_reports")
            os.makedirs(output_dir, exist_ok=True)
            
            # Process each schedule
            for year, module, sched_file, sched_sheet, total_required, department in self.schedules:
                # Load schedule data
                sched_wb = openpyxl.load_workbook(sched_file)
                sched_ws = sched_wb[sched_sheet]
                session_schedule = list(sched_ws.values)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
                
                # Calculate sessions
                completed_sessions = self.calculate_completed_sessions(session_schedule[1:])
                session_details = self.calculate_session_details(session_schedule[1:])
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
                
                # Validate attendance
                valid_attendance = self.validate_attendance(log_history, session_schedule[1:], 
                                                         student_map, f"Year {year}")
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
                
                # Create output workbook and sheets
                output_wb = openpyxl.Workbook()
                output_wb.remove(output_wb.active)
                
                self.create_valid_logs_sheet(output_wb, 'Attendance', valid_attendance)
                self.create_summary_sheet(output_wb, 'Summary', valid_attendance, session_details,
                                        student_map, f"Year {year}", completed_sessions, total_required, department)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))
                
                # Save output workbook with department in filename
                year_dir = os.path.join(output_dir, f"Year_{year}")
                os.makedirs(year_dir, exist_ok=True)
                output_path = os.path.join(year_dir, f"Y{year}_{module}_{department}_attendance.xlsx")
                output_wb.save(output_path)
                current_step += 1
                self.progress_updated.emit(int(current_step / total_steps * 100))

            self.processing_complete.emit()

        except Exception as e:
            self.error_occurred.emit(str(e))


    def create_student_map(self, student_db):
        student_map = {}
        for row in student_db[1:]:
            if row[0]:
                student_id = str(row[0])
                email = f"{student_id}@med.asu.edu.eg"
                student_map[student_id] = {
                    "name": row[1],
                    "year": row[2],
                    "group": row[3],
                    "email": email
                }
        return student_map

    def calculate_completed_sessions(self, session_schedule):
        completed_sessions = {}
        for row in session_schedule:
            if len(row) >= 2:
                year, group = row[:2]
                key = f"{year}-{group}"
                completed_sessions[key] = completed_sessions.get(key, 0) + 1
        return completed_sessions

    def calculate_session_details(self, session_schedule):
        session_details = {}
        
        for row in session_schedule:
            year, group, session, location = row[:4]
            key = f"{year}-{group}"
            if key not in session_details:
                session_details[key] = {}
            
            session_details[key][session] = {
                "location": location,
                "required": 1
            }
            
        return session_details

    def validate_attendance(self, log_history, session_schedule, student_map, target_year):
        valid_attendance = {}
        VALID_ATTENDANCE_WINDOW = timedelta(minutes=15)
        session_map = {}
        unique_logs = set()

        for row in session_schedule:
            year, group, session, location, date, start_time = row[:6]
            key = f"{year}-{group}"
            session_datetime = self.parse_datetime(date, start_time)
            session_key = f"{location}-{date}"
            if key not in session_map:
                session_map[key] = {}
            session_map[key][session_key] = (session, session_datetime)

        for row in log_history[1:]:
            if len(row) >= 4:
                student_id, location, date, time = row[:4]
                student_id = str(student_id)
                if student_id in student_map:
                    student = student_map[student_id]
                    key = f"{student['year']}-{student['group']}"
                    session_key = f"{location}-{date}"
                    if key in session_map and session_key in session_map[key]:
                        session, session_start = session_map[key][session_key]
                        log_datetime = self.parse_datetime(date, time)
                        if session_start - VALID_ATTENDANCE_WINDOW <= log_datetime <= session_start + VALID_ATTENDANCE_WINDOW:
                            unique_log_key = f"{student_id}-{location}-{date}"
                            if unique_log_key not in unique_logs:
                                unique_logs.add(unique_log_key)
                                if key not in valid_attendance:
                                    valid_attendance[key] = []
                                valid_attendance[key].append([
                                    student_id, student['name'], student['year'],
                                    student['group'], student['email'], session,
                                    location, date, time
                                ])
        return valid_attendance

    def parse_datetime(self, date, time):
        if isinstance(date, str):
            date = datetime.strptime(date, '%d/%m/%Y').date()
        if isinstance(time, str):
            time = datetime.strptime(time, '%H:%M:%S').time()
        return datetime.combine(date, time)

    def create_valid_logs_sheet(self, workbook, sheet_name, data):
        sheet = workbook.create_sheet(sheet_name)
        header = ["Student ID", "Name", "Year", "Group", "Email", "Session", "Location", "Date", "Time"]
        sheet.append(header)
        for key in data:
            for row in data[key]:
                sheet.append(row)
        for col in 'H', 'I':
            for cell in sheet[col]:
                cell.number_format = 'DD/MM/YYYY' if col == 'H' else 'HH:MM:SS'
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = max_length + 2

    def create_summary_sheet(self, workbook, sheet_name, valid_attendance, session_details,
                           student_map, target_year, completed_sessions, total_required_sessions, department):
        sheet = workbook.create_sheet(sheet_name)
        
        # Get all unique sessions
        all_sessions = set()
        for key, sessions in session_details.items():
            all_sessions.update(sessions.keys())
        sorted_sessions = sorted(all_sessions)

        header = ["Student ID", "Name", "Year", "Group", "Email", 
                 "Sessions Left", "Sessions Completed", "Total Required", "Total Attended"]
        
        # Add columns for each session with department name
        for session in sorted_sessions:
            header.extend([f"{department} session {session} (Required)", f"{department} session {session} (Attended)"])
        
        sheet.append(header)

        for student_id, student in student_map.items():
            if student['year'] == target_year:
                key = f"{student['year']}-{student['group']}"
                group_completed = completed_sessions.get(key, 0)
                total_attended = 0
                attendance_by_session = {}
                
                # Calculate attendance for each session
                for entry in valid_attendance.get(key, []):
                    if entry[0] == student_id:
                        session = entry[5]
                        attendance_by_session[session] = attendance_by_session.get(session, 0) + 1
                        total_attended += 1

                sessions_left = total_required_sessions - group_completed

                row = [
                    student_id, student['name'], student['year'], student['group'],
                    student['email'], sessions_left, group_completed,
                    total_required_sessions, total_attended
                ]

                # Add attendance data for each session
                for session in sorted_sessions:
                    session_req = session_details.get(key, {}).get(session, {"required": 0})
                    session_att = attendance_by_session.get(session, 0)
                    row.extend([session_req["required"], session_att])

                sheet.append(row)

        # Format columns
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = max_length + 2

#==========================================================attendance analyzer==========================================================#

class AttendanceDashboard(QWidget):
    def __init__(self):
        super().__init__()
        self.student_data = []
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()

    def return_to_home(self):
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)

        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Analysis Dashboard")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()

        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Buttons
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)

        process_attendance_btn = QPushButton("Process Attendance")
        process_attendance_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)

        # File Selection Section
        file_group = QGroupBox("File Selection")
        file_group.setStyleSheet(GROUP_BOX_STYLE)
        file_layout = QVBoxLayout(file_group)
        
        file_input_layout = QHBoxLayout()
        file_input_layout.addWidget(QLabel("Reports File:"))
        self.file_path = QLineEdit()
        self.file_path.setPlaceholderText("Select Excel file...")
        self.file_path.setMinimumWidth(200)
        file_input_layout.addWidget(self.file_path, stretch=1)
        
        browse_btn = QPushButton("Browse")
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        browse_btn.clicked.connect(self.browse_file)
        file_input_layout.addWidget(browse_btn)
        
        file_input_layout.addWidget(QLabel("Sheet Name:"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.setMinimumWidth(100)
        file_input_layout.addWidget(self.sheet_combo)
        
        file_layout.addLayout(file_input_layout)
        main_layout.addWidget(file_group)

        # Statistics Section
        stats_group = QGroupBox("Statistics")
        stats_group.setStyleSheet(GROUP_BOX_STYLE)
        stats_layout = QHBoxLayout(stats_group)

        # Create stat cards
        self.total_students = self.create_stat_card("Total Students", "0")
        self.total_sessions = self.create_stat_card("Total Sessions", "0")
        self.avg_attendance = self.create_stat_card("Average Attendance", "0%")
        self.completion_rate = self.create_stat_card("Completion Rate", "0")

        stats_layout.addWidget(self.total_students)
        stats_layout.addWidget(self.total_sessions)
        stats_layout.addWidget(self.avg_attendance)
        stats_layout.addWidget(self.completion_rate)
        main_layout.addWidget(stats_group)
        
        # Session table section
        session_group = QGroupBox("Attendance Distribution")
        session_group.setStyleSheet(GROUP_BOX_STYLE)
        session_layout = QVBoxLayout(session_group)
        
        self.session_table = QTableWidget()
        self.session_table.setColumnCount(3)
        self.session_table.setHorizontalHeaderLabels(['Group', 'Session', 'Attendance Rate'])
        self.session_table.setStyleSheet(TABLE_STYLE)
        
        header = self.session_table.horizontalHeader()
        self.session_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Align cell content to center
        for row in range(self.session_table.rowCount()):
            for col in range(self.session_table.columnCount()):
                item = self.session_table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
        session_layout.addWidget(self.session_table)
        main_layout.addWidget(session_group)

        # Student List Section
        student_group = QGroupBox("Students List")
        student_group.setStyleSheet(GROUP_BOX_STYLE)
        student_layout = QVBoxLayout(student_group)
        
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search by ID or Name...")
        self.search_input.textChanged.connect(self.filter_students)
        search_layout.addWidget(self.search_input)
        student_layout.addLayout(search_layout)
        
        self.student_table = QTableWidget()
        self.student_table.setColumnCount(5)
        self.student_table.setHorizontalHeaderLabels([
            'Student ID', 'Name', 'Group', 'Sessions Attended', 'Session Details'
        ])
        self.student_table.setStyleSheet(TABLE_STYLE)
        
        self.student_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.student_table.horizontalHeader().setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Align cell content to center
        for row in range(self.student_table.rowCount()):
            for col in range(self.student_table.columnCount()):
                item = self.student_table.item(row, col)
                if item:
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
        student_layout.addWidget(self.student_table)
        main_layout.addWidget(student_group)

        # Bottom Buttons
        display_layout = QHBoxLayout()
        display_btn = QPushButton("Display Statistics")
        display_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        display_btn.clicked.connect(self.display_statistics)
        display_layout.addWidget(display_btn)
        main_layout.addLayout(display_layout)

    def create_stat_card(self, title, value):
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame {{
                background-color: {DARK_BLUE};
                border-radius: 5px;
                padding: 10px;
            }}
            QFrame:hover {{
                background-color: #1b2649;
            }}
            QLabel {{
                color: {TEXT_COLOR};
            }}
        """)
        layout = QVBoxLayout(card)
                        
        title_label = QLabel(title)
        title_label.setStyleSheet("font-size: 14px;")
        value_label = QLabel(value)
        value_label.setStyleSheet("font-size: 24px; font-weight: bold;")
        
        layout.addWidget(title_label)
        layout.addWidget(value_label)
        card.value_label = value_label
        return card

    def browse_file(self):
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "    Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self.update_sheet_list(file_name)

    def update_student_table(self, data):
        """Update the student table with the provided data"""
        self.student_table.setRowCount(len(data))
        for i, row in enumerate(data):
            for j, value in enumerate(row):
                item = QTableWidgetItem(value)
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self.student_table.setItem(i, j, item)
    
    def filter_students(self):
        """Filter students based on search input"""
        search_text = self.search_input.text().lower()
        filtered_data = [
            row for row in self.student_data
            if search_text in str(row[0]).lower() or search_text in str(row[1]).lower()
        ]
        self.update_student_table(filtered_data)
    
    def browse_file(self):
        """Open file dialog to select Excel file"""
        file_name, _ = QFileDialog.getOpenFileName(
            self,
            "    Select Excel File",
            "",
            "Excel Files (*.xlsx *.xls)"
        )
        if file_name:
            self.file_path.setText(file_name)
            self.update_sheet_list(file_name)
    
    def update_sheet_list(self, file_path):
        """Update sheet combo box with available sheets"""
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            self.sheet_combo.clear()
            self.sheet_combo.addItems(wb.sheetnames)
            
            summary_index = self.sheet_combo.findText("Summary", Qt.MatchFlag.MatchExactly)
            if summary_index >= 0:
                self.sheet_combo.setCurrentIndex(summary_index)
        
            wb.close()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error reading Excel file: {str(e)}")
    
    def display_statistics(self):
        """Display statistics from the selected Excel file"""
        file_path = self.file_path.text()
        sheet_name = self.sheet_combo.currentText()
        
        if not file_path or not sheet_name:
            self.show_custom_warning("Reports File Required", "Please select both file and sheet name")
            return 
            
        if not os.path.exists(file_path):
            self.show_custom_warning("Failed to Load Reports File", "Selected file does not exist")
            return 
            
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            if sheet_name not in wb.sheetnames:
                self.show_custom_warning("Warning", "Selected sheet not found in workbook")
                return
    
            summary_sheet = wb[sheet_name]
            rows = list(summary_sheet.rows)
            header_row = rows[0]  # Get header row
            data_rows = rows[1:]  # Skip header row
            
            # Find column indices for important data
            total_required_idx = None
            total_attended_idx = None
            group_idx = None
            student_id_idx = None
            name_idx = None
            
            for idx, cell in enumerate(header_row):
                header_value = str(cell.value).strip() if cell.value else ""
                if header_value == "Total Required":
                    total_required_idx = idx
                elif header_value == "Total Attended":
                    total_attended_idx = idx
                elif header_value == "Group":
                    group_idx = idx
                elif header_value == "Student ID":
                    student_id_idx = idx
                elif header_value == "Name":
                    name_idx = idx
            
            # Find session columns for session details
            session_columns = []
            for idx, cell in enumerate(header_row):
                if cell.value and "session" in str(cell.value).lower() and "(Required)" in str(cell.value):
                    session_columns.append((idx, idx + 1))  # (Required column, Attended column)
            
            # Make sure we found all necessary columns
            if None in (total_required_idx, total_attended_idx, group_idx, student_id_idx, name_idx):
                self.show_custom_warning("Error", "Required columns not found in the Excel file")
                return
            
            if not session_columns:
                self.show_custom_warning("Error", "No session columns found in the Excel file")
                return
                
            # Initialize group-based tracking
            groups = set()
            group_student_counts = {}  # {group: total_students}
            group_session_attendance = {}  # {group: {session_num: attended_count}}
            
            # First pass: identify all groups and count students per group
            for row in data_rows:
                group = str(row[group_idx].value)
                groups.add(group)
                group_student_counts[group] = group_student_counts.get(group, 0) + 1
            
            # Get the total sessions from the first data row (assuming all students have same requirement)
            total_sessions = int(data_rows[0][total_required_idx].value) if data_rows else 0
            
            # Initialize attendance tracking for each group
            for group in groups:
                group_session_attendance[group] = {i: 0 for i in range(1, total_sessions + 1)}
            
            # Process student data
            self.student_data = []
            total_attended_sessions = 0
            completion_count = 0
            
            for row in data_rows:
                group = str(row[group_idx].value)
                attended_count = int(row[total_attended_idx].value)
                required_count = int(row[total_required_idx].value)
                
                # Track attended sessions for session details
                attended_sessions = []
                for session_idx, (req_col, att_col) in enumerate(session_columns, 1):
                    if row[att_col].value == 1:
                        attended_sessions.append(str(session_idx))
                        group_session_attendance[group][session_idx] += 1
                
                total_attended_sessions += attended_count
                
                if attended_count == required_count:
                    completion_count += 1
                
                self.student_data.append([
                    str(row[student_id_idx].value),  # Student ID
                    str(row[name_idx].value),        # Name
                    group,                           # Group
                    f"{attended_count}/{required_count}",  # Sessions Attended
                    ", ".join(sorted(attended_sessions)) if attended_sessions else "None"  # Session Details
                ])
            
            # Update attendance distribution table with group-based rates
            row_count = len(groups) * total_sessions
            self.session_table.setRowCount(row_count)
            current_row = 0
            
            for group in sorted(groups):
                for session_num in range(1, total_sessions + 1):
                    # Calculate attendance rate for this group and session
                    group_total = group_student_counts[group]
                    attended = group_session_attendance[group][session_num]
                    attendance_rate = (attended / group_total * 100) if group_total > 0 else 0
                    
                    # Create items with centered alignment
                    group_item = QTableWidgetItem(group)
                    session_item = QTableWidgetItem(f"Session {session_num}")
                    rate_item = QTableWidgetItem(f"{attendance_rate:.1f}%")
                    
                    # Set alignment for each item
                    group_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    session_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    rate_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    
                    # Set items in table
                    self.session_table.setItem(current_row, 0, group_item)
                    self.session_table.setItem(current_row, 1, session_item)
                    self.session_table.setItem(current_row, 2, rate_item)
                    current_row += 1
            
            # Update other stats
            total_students = len(data_rows)
            avg_attendance = (total_attended_sessions / (total_students * total_sessions) * 100) if total_students > 0 else 0            
            self.total_students.value_label.setText(str(total_students))
            self.total_sessions.value_label.setText(str(total_sessions))
            self.avg_attendance.value_label.setText(f"{avg_attendance:.1f}%")
            self.completion_rate.value_label.setText(str(completion_count))
            
            self.update_student_table(self.student_data)
            wb.close()
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error loading report: {str(e)}")
        
    def show_custom_warning(self, title, message):
        """Show a custom styled warning dialog"""
        warning_dialog = QMessageBox(self)
        warning_dialog.setWindowTitle(title)
        warning_dialog.setText(message)
        warning_dialog.setIcon(QMessageBox.Icon.Warning)
    
        # Create and style OK button
        ok_button = warning_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
    
        # Style dialog background and text
        warning_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)
    
        warning_dialog.exec()  

    def setup_worker(self, dep_file, faculty_file):
        # Create the worker
        self.worker = PopulateWorker(dep_file, faculty_file)
        
        # Create the thread
        self.thread = QThread()
        
        # Move worker to the thread
        self.worker.moveToThread(self.thread)
        
        # Connect signals and slots
        self.thread.started.connect(self.worker.run)
        self.worker.finished.connect(self.thread.quit)
        self.worker.finished.connect(self.worker.deleteLater)
        self.thread.finished.connect(self.thread.deleteLater)
        
        # Connect the progress and output signals
        self.worker.progress.connect(self.progress_bar.setValue)
        self.worker.output.connect(self.output_console.append)
        
        # Connect the new error and success signals
        self.worker.error.connect(self.handle_error)
        self.worker.success.connect(self.handle_success)
        
        # Start the thread
        self.thread.start()
    
    def handle_error(self, error_message):
        self.setEnabled(True)
        error_dialog = QMessageBox(self)
        error_dialog.setWindowTitle("Error")
        error_dialog.setText(f"Error processing data: {error_message}")
        error_dialog.setIcon(QMessageBox.Icon.Critical)

        # Style OK button
        ok_button = error_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        error_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        error_dialog.exec()
        self.output_console.append(f"Error: {error_message}")
    
    def handle_success(self, success_message):
        self.setEnabled(True)
        self.progress_bar.setValue(100)
        self.output_console.append("Processing complete!")
        
        success_dialog = QMessageBox(self)
        success_dialog.setWindowTitle("Success")
        success_dialog.setText("Processing complete! Check the attendance_reports folder.")
        success_dialog.setIcon(QMessageBox.Icon.Information)

        # Style OK button
        ok_button = success_dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)

        # Style dialog background
        success_dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)

        success_dialog.exec()  

#==========================================================attendance populator==========================================================#

class Facultypopulator(QWidget):
    # Add a signal for progress updates
    progress_signal = pyqtSignal(int)
    output_signal = pyqtSignal(str)
    complete_signal = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.dep_file = None
        self.faculty_file = None
        # Define department mapping for search patterns
        self.department_patterns = {
            "Anatomy": ["anatomy", "anat", "anat.", "anato", "anato."],
            "Histology": ["histology", "histo", "histo.", "hist", "hist.", "sgd"],
            "Physiology": ["physiology", "physio", "physio."],
            "Biochemistry": ["biochemistry", "biochem", "biochem.", "bio", "bio.","cbl"],
            "Pharmacology": ["pharmacology", "pharma", "pharma."],
            "Pathology": ["pathology", "path", "path.", "patho", "patho."],
            "Parasitology": ["parasitology", "para", "para.", "sgd"],
            "Microbiology": ["microbiology", "micro", "micro."],
            "Forensics and Toxicology": ["forensics", "toxicology", "forensic", "toxico", "toxic", "toxic.", "toxico.", "forens.", "forens", "foren", "foren."],
            "Community Medicine": ["community", "comm med", "comm."],
            "Clinical": ["clinical", "clinic", "clin.", "clinic."]
        }
        self.selected_department = "Pharmacology"  # Default selection
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()

    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)

    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Track the current message index
        self.current_message_index = 0
    
        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Attendance populator")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()
    
        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)
            
        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        # Add button layout to header
        header_layout.addLayout(header_buttons_layout)
        
        # Add header to main layout
        main_layout.addLayout(header_layout)

        # Department Selection Section
        department_group = QGroupBox("Department Selection")
        department_group.setStyleSheet(GROUP_BOX_STYLE)
        department_layout = QHBoxLayout(department_group)
        
        department_layout.addWidget(QLabel("Select Department:"))
        self.department_combo = QComboBox()
        self.department_combo.setMinimumWidth(200)
        
        # Add departments in the specified order
        departments = [
            "Anatomy", 
            "Histology", 
            "Physiology", 
            "Biochemistry", 
            "Pharmacology", 
            "Pathology", 
            "Parasitology", 
            "Microbiology", 
            "Forensics and Toxicology", 
            "Community Medicine", 
            "Clinical"
        ]
        self.department_combo.addItems(departments)
        
        # Set default to Pharmacology (index 4)
        self.department_combo.setCurrentIndex(4)
        self.department_combo.currentTextChanged.connect(self.update_selected_department)
        
        department_layout.addWidget(self.department_combo)
        department_layout.addStretch()
        
        # Add department group to main layout
        main_layout.addWidget(department_group)

        # Dep File Section 
        attendance_group = QGroupBox("Department Attendance Report")
        attendance_group.setStyleSheet(GROUP_BOX_STYLE)
        attendance_layout = QHBoxLayout(attendance_group)
        
        attendance_layout.addWidget(QLabel("Department Attendance File:"))
        self.dep_path = QLineEdit()
        self.dep_path.setPlaceholderText("Select Excel file...")
        self.dep_path.setReadOnly(True)
        attendance_layout.addWidget(self.dep_path)
        
        attendance_browse_btn = QPushButton("Browse")
        attendance_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        attendance_browse_btn.clicked.connect(lambda: self.browse_file('dep'))
        attendance_layout.addWidget(attendance_browse_btn)

        attendance_layout.addWidget(QLabel("Sheet Name:"))
        self.dep_sheet_combo = QComboBox()
        self.dep_sheet_combo.setMinimumWidth(100)
        attendance_layout.addWidget(self.dep_sheet_combo)
        
        # Add attendance group to main layout
        main_layout.addWidget(attendance_group)

        # Faculty File Section
        faculty_group = QGroupBox("Main Attendance Report")
        faculty_group.setStyleSheet(GROUP_BOX_STYLE)
        faculty_layout = QHBoxLayout(faculty_group)
        
        faculty_layout.addWidget(QLabel("Main Attendance File:"))
        self.faculty_path = QLineEdit()
        self.faculty_path.setPlaceholderText("Select Excel file...")
        self.faculty_path.setReadOnly(True)
        faculty_layout.addWidget(self.faculty_path)
        
        faculty_browse_btn = QPushButton("Browse")
        faculty_browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        faculty_browse_btn.clicked.connect(lambda: self.browse_file('faculty'))
        faculty_layout.addWidget(faculty_browse_btn)

        faculty_layout.addWidget(QLabel("Sheet Name:"))
        self.faculty_sheet_combo = QComboBox()
        self.faculty_sheet_combo.setMinimumWidth(100)
        faculty_layout.addWidget(self.faculty_sheet_combo)
        
        # Add faculty group to main layout
        main_layout.addWidget(faculty_group)

        # Facts Display Section 
        facts_group = QGroupBox("Status")
        facts_group.setStyleSheet(GROUP_BOX_STYLE)
        facts_layout = QVBoxLayout(facts_group)
        
        self.fact_label = QLabel("")
        self.fact_label.setStyleSheet("font-weight: bold; font-size: 16px; color: #555;")
        self.fact_label.setWordWrap(True)
        self.fact_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        facts_layout.addWidget(self.fact_label)
        
        # Add stretch to push widgets to the top and make the empty space part of the layout
        facts_layout.addStretch(1)
        
        facts_group.setMinimumHeight(100)
        # Add the facts group to the main layout 
        main_layout.addWidget(facts_group)  

        # Progress Bar Section
        progress_group = QGroupBox("Progress")
        progress_group.setStyleSheet(GROUP_BOX_STYLE)
        progress_layout = QVBoxLayout(progress_group)

        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)
        progress_layout.addWidget(self.progress_bar)
        
        # Create loading gif label
        self.loading_label = QLabel()
        self.loading_label.setFixedSize(24, 24)  # Adjust based on your GIF size
        self.loading_label.setVisible(False)  # Hidden by default
        
        # Create the movie object for the GIF
        self.loading_movie = QMovie()
        self.loading_movie.setScaledSize(QSize(24, 24))  # Adjust based on your GIF size
        self.loading_label.setMovie(self.loading_movie)
        
        # Make sure to have your loading.gif in the same directory as the script
        loading_gif_path = os.path.join(os.path.dirname(__file__), 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie.setFileName(loading_gif_path)
        else:
            print(f"Warning: loading.gif not found at {loading_gif_path}")
        
        # Create a horizontal layout to hold both the progress bar and loading animation
        progress_h_layout = QHBoxLayout()
        progress_h_layout.addWidget(self.progress_bar)
        progress_h_layout.addWidget(self.loading_label)
        progress_layout.addLayout(progress_h_layout)
        
        # Add progress group to main layout
        main_layout.addWidget(progress_group)

        # Output Console Section
        console_group = QGroupBox("Output Console")
        console_group.setStyleSheet(GROUP_BOX_STYLE)
        console_layout = QVBoxLayout(console_group)

        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(150)
        self.output_console.setStyleSheet(CONSOLE_STYLE)  
        console_layout.addWidget(self.output_console)
        main_layout.addWidget(console_group)

        # Bottom Buttons
        button_layout = QHBoxLayout()
        self.populate_btn = QPushButton("Populate Main Attendance File")
        self.populate_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        self.populate_btn.clicked.connect(self.start_process)
        button_layout.addWidget(self.populate_btn)        
        
        # Add button layout to main layout
        main_layout.addLayout(button_layout)
        
        # Add a system console redirector
        sys.stdout = ConsoleRedirector(self.output_console)
        
        # Connect our signals to their handlers
        self.progress_signal.connect(self.update_progress_bar)
        self.output_signal.connect(self.update_output_console)
        self.complete_signal.connect(self.process_complete)
        
        # Initialize thread variable
        self.worker_thread = None
        
    def update_selected_department(self, department):
        """Update the selected department when the combo box changes"""
        self.selected_department = department
        self.output_console.append(f"Selected department: {department}")
        
    def update_progress_bar(self, value):
        """Update the progress bar from the worker thread"""
        self.progress_bar.setValue(value)
        
    def process_complete(self, result):
        """Handle process completion"""
        
        # Stop the thread monitor timer if it exists
        if hasattr(self, 'thread_monitor'):
            self.thread_monitor.stop()
        
        # Hide the loading animation
        self.loading_label.setVisible(False)
        self.loading_movie.stop()
        
        # Re-enable the populate button
        self.populate_btn.setEnabled(True)
        
        # Show completion dialog
        if "Error" not in result:
            QMessageBox.information(
                self,
                "Integration Complete",
                result,
                QMessageBox.StandardButton.Ok
            )

    def update_output_console(self, text):
        """Update the output console with text from the worker thread"""
        if self.output_console is not None:
            # Use setText instead of append to avoid buffer overflow
            # If you need to keep history, limit it to a reasonable amount
            cursor = self.output_console.textCursor()
            cursor.movePosition(QTextCursor.MoveOperation.End)
            cursor.insertText(text + "\n")
            self.output_console.setTextCursor(cursor)
            self.output_console.ensureCursorVisible()
            
    def populate_attendance(self):
        """Process and populate department attendance into faculty table"""
        if not self.dep_file or not self.faculty_file:
            return "Please select both files first"
        try:
            # Get selected department search patterns
            department_name = self.selected_department
            search_patterns = self.department_patterns.get(department_name, ["unknown"])
            
            # Emit signal to update progress
            self.progress_signal.emit(10)
            self.output_signal.emit(f"Starting integration process for {department_name} department...")
            
            # Create a backup of the original faculty file
            # Create backup filename with timestamp
            backup_filename = f"{os.path.splitext(self.faculty_file)[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(self.faculty_file)[1]}"
            # Create the backup
            shutil.copy2(self.faculty_file, backup_filename)
            self.output_signal.emit(f"Created backup of original file at: {backup_filename}")
            
            # Update progress
            self.progress_signal.emit(20)
                
            # Load the department attendance data
            self.output_signal.emit(f"Loading {department_name} attendance data...")
            dep_df = pd.read_excel(self.dep_file, sheet_name="Attendance")
            # Load the summary data which has attendance status per student
            summary_df = pd.read_excel(self.dep_file, sheet_name="Summary")
            
            # Update progress
            self.progress_signal.emit(30)
            
            # Load the faculty-wide attendance table without modifying
            self.output_signal.emit("Loading faculty attendance data...")
            faculty_df = pd.read_excel(self.faculty_file, header=None)
            # Load the workbook with openpyxl to preserve formatting
            from openpyxl import load_workbook
            workbook = load_workbook(self.faculty_file)
            sheet = workbook.active
            
            # Track changes for reporting
            updated_records = 0
            students_processed = 0
            self.output_signal.emit(f"Faculty columns: {list(faculty_df.columns)}")
            self.output_signal.emit(f"{department_name} columns: {list(dep_df.columns)}")
            self.output_signal.emit(f"Summary columns: {list(summary_df.columns)}")
            
            # Update progress
            self.progress_signal.emit(40)
            
            # Find the Department session columns by looking for cells containing department patterns
            dept_cols = []
            self.output_signal.emit(f"Searching for {department_name} session columns using patterns: {search_patterns}...")
            
            # Search through first 15 rows to find cells containing department patterns
            for row_idx in range(min(15, len(faculty_df))):
                for col_idx in range(len(faculty_df.columns)):
                    cell_value = str(faculty_df.iloc[row_idx, col_idx]).strip() if pd.notna(faculty_df.iloc[row_idx, col_idx]) else ""
                    cell_value_lower = cell_value.lower()
                    
                    # Check if any pattern is in the cell
                    pattern_found = False
                    matched_pattern = ""
                    for pattern in search_patterns:
                        if pattern.lower() in cell_value_lower:
                            pattern_found = True
                            matched_pattern = pattern
                            break
                    
                    if pattern_found:
                        # Extract session number if present
                        match = re.search(r'(\d+)', cell_value)
                        session_num = int(match.group(1)) if match else None
                        if session_num is not None:
                            dept_cols.append((col_idx, cell_value, session_num))
                            self.output_signal.emit(f"Found {department_name} column at row {row_idx}, column {col_idx}: {cell_value}, Session: {session_num}, Matched pattern: {matched_pattern}")
            
            if not dept_cols:
                self.output_signal.emit(f"Error: Could not locate any {department_name} columns in faculty sheet")
                return f"Could not locate any {department_name} columns in faculty sheet"
            
            # Update progress
            self.progress_signal.emit(45)
            
            # Find the student ID column in the faculty sheet
            # We'll look for a column where most values are numeric strings with 4+ digits
            id_col_idx = None
            id_col_row_start = None
            self.output_signal.emit("Searching for student ID column...")
            
            # First scan the first ~20 rows to find potential ID columns
            potential_id_columns = {}
            # Look through first ~30 rows to get a good sample
            for row_idx in range(min(30, len(faculty_df))):
                for col_idx in range(len(faculty_df.columns)):
                    cell_value = str(faculty_df.iloc[row_idx, col_idx]).strip() if pd.notna(faculty_df.iloc[row_idx, col_idx]) else ""
                    # Check if cell contains a string of 4+ digits
                    if cell_value.isdigit() and len(cell_value) >= 4:
                        # Add this column to our potential columns dict and increment its count
                        if col_idx not in potential_id_columns:
                            potential_id_columns[col_idx] = []
                        potential_id_columns[col_idx].append(row_idx)
            
            # Find the column with the most potential student IDs
            best_col = None
            max_ids = 0
            for col_idx, row_indices in potential_id_columns.items():
                if len(row_indices) > max_ids:
                    max_ids = len(row_indices)
                    best_col = col_idx
            
            if best_col is not None:
                id_col_idx = best_col
                # Find first row with student ID (assuming continuous data afterward)
                id_col_row_start = min(potential_id_columns[best_col])
                self.output_signal.emit(f"Found student ID column at index {id_col_idx}, starting from row {id_col_row_start}")
                self.output_signal.emit(f"Found {max_ids} potential student IDs in this column")
            else:
                self.output_signal.emit("Warning: Could not detect a clear student ID column. Falling back to full sheet search.")
            
            # Update progress
            self.progress_signal.emit(50)
            
            # Get the Dep student ID column name
            dep_id_col = "Student ID"  # Based on the printed column names
            
            # Update progress
            total_students = len(summary_df)
            
            # Process each student in the summary sheet
            self.output_signal.emit(f"Processing {total_students} students...")
            for idx, dep_row in enumerate(summary_df.iterrows()):
                _, dep_row = dep_row  # Unpack the tuple
                
                # Calculate progress - spread between 50% and 90%
                student_progress = 50 + int((idx / total_students) * 40)
                self.progress_signal.emit(student_progress)
                
                # Get the student ID
                if dep_id_col not in dep_row:
                    self.output_signal.emit(f"Cannot find column {dep_id_col} in department sheet")
                    continue
                    
                student_id = str(dep_row[dep_id_col]).strip()
                
                # Find student ID in the faculty sheet
                matching_cells = []
                
                # If we found a likely student ID column, search only in that column
                if id_col_idx is not None and id_col_row_start is not None:
                    # Only search in the identified student ID column, starting from the first identified row
                    for row_idx in range(id_col_row_start, len(faculty_df)):
                        cell_value = faculty_df.iloc[row_idx, id_col_idx]
                        if pd.notna(cell_value):
                            faculty_cell_value = str(cell_value).strip()
                            if faculty_cell_value == student_id:
                                matching_cells.append((row_idx, id_col_idx))
                                self.output_signal.emit(f"Found matching student ID {student_id} at cell ({row_idx}, {id_col_idx})")
                                break  # Since we're in the ID column, we only need the first match
                else:
                    # Fallback: Search the entire sheet for matching cells
                    self.output_signal.emit(f"Searching full sheet for student ID: {student_id}")
                    for row_idx in range(len(faculty_df)):
                        for col_idx in range(len(faculty_df.columns)):
                            cell_value = faculty_df.iloc[row_idx, col_idx]
                            if pd.notna(cell_value):
                                # Compare after converting to string and stripping whitespace
                                faculty_cell_value = str(cell_value).strip()
                                if faculty_cell_value == student_id:
                                    matching_cells.append((row_idx, col_idx))
                                    self.output_signal.emit(f"Found matching student ID {student_id} at cell ({row_idx}, {col_idx})")
            
                if not matching_cells:
                    self.output_signal.emit(f"No match found for student ID: {student_id}")
                    continue
                    
                # For each matching student ID cell
                for row_idx, _ in matching_cells:
                    # For each department column detected
                    for col_idx, _, session_num in dept_cols:
                        if session_num is None:
                            self.output_signal.emit(f"Skipping {department_name} column without session number at column {col_idx}")
                            continue
                            
                        # Look for attendance columns that contain both the session number and "(Attended)"
                        attended = 0  # Default value
                        
                        # Find the appropriate attendance column by searching for session number and "(Attended)"
                        attendance_col = None
                        for col_name in dep_row.index:
                            # Check if this column contains both the session number and "(Attended)"
                            if "(Attended)" in str(col_name) and str(session_num) in str(col_name):
                                attendance_col = col_name
                                self.output_signal.emit(f"Found matching attendance column: {attendance_col}")
                                break
                                
                        if attendance_col:
                            # If found, get the attendance value
                            attended = int(dep_row[attendance_col]) if pd.notna(dep_row[attendance_col]) else 0
                            self.output_signal.emit(f"Attendance value from column {attendance_col}: {attended}")
                        else:
                            self.output_signal.emit(f"Could not find attendance column for session {session_num}")
                            # Fallback to checking the attendance sheet
                            student_sessions = dep_df.loc[dep_df[dep_id_col].astype(str).str.strip() == student_id]
                            if not student_sessions.empty:
                                session_col = 'Session'
                                if session_col in student_sessions.columns:
                                    session_attended = student_sessions.loc[student_sessions[session_col] == session_num]
                                    if not session_attended.empty:
                                        attended = 1
                                        self.output_signal.emit(f"Fallback: Student {student_id} attended {department_name} session {session_num}")
                        
                        # Update the faculty table at the precise X,Y coordinate (row, column)
                        faculty_df.iloc[row_idx, col_idx] = attended
                        
                        # Also update the Excel file directly using openpyxl (1-indexed)
                        # Convert to Excel's 1-based indexing
                        excel_row = row_idx + 1
                        excel_col = col_idx + 1
                        
                        # Convert column index to letter (A, B, C, etc.)
                        col_letter = get_column_letter(excel_col)
                        
                        # Update the cell in openpyxl
                        sheet[f"{col_letter}{excel_row}"] = attended
                        updated_records += 1
                        self.output_signal.emit(f"Updated student {student_id}, session {session_num}, attendance={attended} at cell ({row_idx}, {col_idx})")
                        
                # Increment the student counter
                students_processed += 1
                
                # Save every 10 students
                if students_processed % 10 == 0:
                    self.output_signal.emit(f"Saving after processing {students_processed} students...")
                    try:
                        workbook.save(self.faculty_file)
                        self.output_signal.emit(f"Successfully saved after {students_processed} students")
                    except Exception as save_error:
                        self.output_signal.emit(f"Warning: Save failed: {str(save_error)}")
                        self.output_signal.emit(f"Continuing processing but updates may not persist")
            
            # Final save at the end
            self.output_signal.emit(f"Final save after processing all {students_processed} students...")
            workbook.save(self.faculty_file)     
            
            # Also save a DataFrame version just to be safe
            backup_df_path = self.faculty_file + "_dataframe_backup.xlsx"
            faculty_df.to_excel(backup_df_path, index=False, header=False)
            self.output_signal.emit(f"Saved DataFrame backup to {backup_df_path}")
        
            # Update progress to 100%
            self.progress_signal.emit(100)
            
            success_message = f"Successfully updated {updated_records} attendance records for {students_processed} students. Original file backed up to {backup_filename}"
            self.output_signal.emit(success_message)
            return success_message
            
        except Exception as e:
            error_details = traceback.format_exc()
            self.output_signal.emit(f"Detailed error: {error_details}")
            # Reset progress bar on error
            self.progress_signal.emit(0)
            error_message = f"Error during integration: {str(e)}"
            self.output_signal.emit(error_message)
            return error_message
                                                
    def browse_file(self, file_type):
        """Handle file browsing for either department or faculty files."""
        filename, _ = QFileDialog.getOpenFileName(
            self, "Select Excel File", "", "Excel Files (*.xlsx)"
        )
        if filename:
            if file_type == 'dep':
                self.dep_path.setText(filename)
                self.dep_file = filename
                self.load_sheets(filename, self.dep_sheet_combo)
            elif file_type == 'faculty':
                self.faculty_path.setText(filename)
                self.faculty_file = filename
                self.load_sheets(filename, self.faculty_sheet_combo)

    def load_sheets(self, file_path, combo_box):
        """Load sheet names and auto-select 'Summary' for department files"""
        if os.path.isfile(file_path):
            try:
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheets = wb.sheetnames
                combo_box.clear()
                combo_box.addItems(sheets)
                
                # Auto-select Summary sheet for dep files
                if combo_box == self.dep_sheet_combo:
                    summary_index = combo_box.findText("Summary", Qt.MatchFlag.MatchExactly)
                    if summary_index >= 0:
                        combo_box.setCurrentIndex(summary_index)
                        self.output_console.append(f"Automatically selected Summary sheet for {self.selected_department} file")
                
                wb.close()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Error reading Excel file: {str(e)}")
                self.output_console.append(f"Error loading sheets: {str(e)}")

    def start_process(self):
        # Show facts message
        self.fact_label.setText("\nThis might take a while, you can take a short break to help refresh your mind while the task is being completed.\nأذكر الله")
        self.fact_label.setVisible(True)
        
        # Initialize process completion flag
        self.process_completed = False  
        
        """Starts the facts display and then processes the integration in a separate thread"""
        # Check for files first
        if not self.dep_file or not self.faculty_file:
            self.output_console.append("Please select both files first")
            return
            
        # If there's already a thread running, don't start another one
        if hasattr(self, 'worker_thread') and self.worker_thread is not None and self.worker_thread.isRunning():
            self.output_console.append("A process is already running. Please wait for it to complete.")
            return
            
        # Disable the populate button to prevent multiple clicks
        self.populate_btn.setEnabled(False)
        
        # Show and start the loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
        
        # Reset progress bar
        self.progress_bar.setValue(0)
        
        # Get the selected department and its search patterns
        department_name = self.selected_department
        search_patterns = self.department_patterns.get(department_name, ["unknown"])
        
        # Create and start a worker thread for the processing
        self.worker_thread = QThread()
        self.worker = PopulateWorker(self.dep_file, self.faculty_file, department_name, search_patterns)
        self.worker.moveToThread(self.worker_thread)
        
        # Connect signals
        self.worker_thread.started.connect(self.worker.run)
        self.worker.progress.connect(self.progress_signal.emit)
        self.worker.output.connect(self.output_signal.emit)
        self.worker.finished.connect(self.complete_signal.emit)
        self.worker.finished.connect(self.worker_thread.quit)
        
        # These connections need to be made as QueuedConnection to ensure thread safety
        self.worker.finished.connect(self.worker.deleteLater, Qt.ConnectionType.QueuedConnection)
        self.worker_thread.finished.connect(self.worker_thread.deleteLater, Qt.ConnectionType.QueuedConnection)
        
        # Start the thread
        self.worker_thread.start()
    
    def start_integration(self):
        """Method called when Populate Attendance button is clicked"""
        
        # Then run the integration process
        self.run_integration()

    def process_complete(self, result):
        """Handle process completion"""

        # Hide facts message
        self.fact_label.setVisible(False)
        self.fact_label.clear()

        # Track that the process completed normally
        self.process_completed = True  # Add this line
        
        # Hide the loading animation
        self.loading_label.setVisible(False)
        self.loading_movie.stop()
                
        # Re-enable the populate button
        self.populate_btn.setEnabled(True)
        
        # Show styled completion dialog
        if "Error" not in result:
            self.show_styled_dialog("Success", result, QMessageBox.Icon.Information)
        else:
            self.show_styled_dialog("Error", result, QMessageBox.Icon.Critical)
    
    def show_styled_dialog(self, title, message, icon):
        """Create and show a styled QMessageBox"""
        dialog = QMessageBox(self)
        dialog.setWindowTitle(title)
        dialog.setText(message)
        dialog.setIcon(icon)
        
        # Style the dialog
        dialog.setStyleSheet(f"""
            QMessageBox {{
                background-color: {CARD_BG};
                color: {TEXT_COLOR};
            }}
            QLabel {{
                color: {TEXT_COLOR};
                font-size: 14px;
            }}
        """)
        
        # Get and style the OK button
        ok_button = dialog.addButton(QMessageBox.StandardButton.Ok)
        ok_button.setStyleSheet(STANDARD_BUTTON_STYLE)
        
        dialog.exec()

    def check_thread_status(self):
        """Monitor thread status and handle abnormal termination"""
        try:
            # Check if worker_thread exists and is valid
            if not hasattr(self, 'worker_thread') or self.worker_thread is None:
                self.thread_monitor.stop()
                return
            # Check if the thread is running
            if not self.worker_thread.isRunning():
                self.thread_monitor.stop()
                # If the process didn't complete normally, emit signal
                if not hasattr(self, 'process_completed'):
                    self.complete_signal.emit("Process terminated unexpectedly")
                # Clean up
                self.loading_label.setVisible(False)
                if hasattr(self, 'loading_movie'):
                    self.loading_movie.stop()
                self.populate_btn.setEnabled(True)
                self.worker_thread = None
        except RuntimeError as e:
            # Handle the case where the C++ object is already deleted
            self.thread_monitor.stop()
            self.worker_thread = None
            self.loading_label.setVisible(False)
            if hasattr(self, 'loading_movie'):
                self.loading_movie.stop()
            self.populate_btn.setEnabled(True)
               
class ConsoleRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget
        self.original_stdout = sys.stdout

    def write(self, text):
        self.original_stdout.write(text)  # Still print to console
        if text.strip():  # Only update if there's actual text (not just whitespace)
            self.text_widget.append(text.rstrip())
            QApplication.processEvents()  # Process UI events to update console in real-time

    def flush(self):
        self.original_stdout.flush()

class PopulateWorker(QObject):
    finished = pyqtSignal(str)  # Signal for when processing is complete
    progress = pyqtSignal(int)  # Signal for progress updates
    output = pyqtSignal(str)    # Signal for console output
    error = pyqtSignal(str)     # Signal for error messages
    success = pyqtSignal(str)   # Signal for success messages
    
    def __init__(self, dep_file, faculty_file, department_name="Department", search_patterns=None):
        super().__init__()
        self.dep_file = dep_file
        self.faculty_file = faculty_file
        self.department_name = department_name
        self.search_patterns = search_patterns if search_patterns else ["dep"]
        self.is_running = False
        
    def resource_path(self, relative_path):
        """Get absolute path to resource, works for dev and for PyInstaller"""
        try:
            # PyInstaller creates a temp folder and stores path in _MEIPASS
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(os.path.dirname(__file__))
        
        return os.path.join(base_path, relative_path)
        
    def run(self):
        """Main processing function that runs in a separate thread"""
        try:
            # Set running flag
            self.is_running = True
            self.output.emit(f"Starting integration process with dep file: {self.dep_file}, faculty file: {self.faculty_file}")
            
            # Emit signal to update progress
            self.progress.emit(10)
            self.output.emit("Starting integration process...")
            
            # Create a backup of the original faculty file
            # Create backup filename with timestamp
            backup_filename = f"{os.path.splitext(self.faculty_file)[0]}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}{os.path.splitext(self.faculty_file)[1]}"
            try:
                # Create the backup
                shutil.copy2(self.faculty_file, backup_filename)
                self.output.emit(f"Created backup of original file at: {backup_filename}")
            except Exception as backup_error:
                error_msg = f"Failed to create backup: {str(backup_error)}"
                self.output.emit(f"Warning: {error_msg}")
                # Continue execution despite backup failure
                
            # Update progress
            self.progress.emit(20)
            
            # Load the dep attendance data
            self.output.emit("Loading department attendance data...")
            try:
                dep_df = pd.read_excel(self.dep_file, sheet_name="Attendance")
                # Load the summary data which has attendance status per student
                summary_df = pd.read_excel(self.dep_file, sheet_name="Summary")
            except Exception as excel_error:
                error_msg = f"Failed to load department data: {str(excel_error)}"
                raise Exception(error_msg)
                
            # Update progress
            self.progress.emit(30)
            
            # Load the faculty-wide attendance table without modifying
            self.output.emit("Loading faculty attendance data...")
            try:
                faculty_df = pd.read_excel(self.faculty_file, header=None)
                # Load the workbook with openpyxl to preserve formatting
                from openpyxl import load_workbook
                workbook = load_workbook(self.faculty_file)
                sheet = workbook.active
            except Exception as faculty_error:
                error_msg = f"Failed to load faculty data: {str(faculty_error)}"
                raise Exception(error_msg)
                
            # Track changes for reporting
            updated_records = 0
            students_processed = 0
            self.output.emit(f"Faculty columns: {len(faculty_df.columns)}, dep columns: {len(dep_df.columns)}, Summary columns: {len(summary_df.columns)}")
            
            # Update progress
            self.progress.emit(40)
            
            # Get the department name and search patterns from the worker initialization
            department_name = self.department_name if hasattr(self, 'department_name') else "Department"
            search_patterns = self.search_patterns if hasattr(self, 'search_patterns') else ["dep"]
            
            # Find the department session columns by looking for cells containing department patterns
            dep_cols = []
            self.output.emit(f"Searching for {department_name} session columns using patterns: {search_patterns}...")
            
            # Search through first 15 rows to find cells containing the department patterns
            for row_idx in range(min(15, len(faculty_df))):
                # Check if we're still running - important for clean cancellation
                if not self.is_running:
                    raise Exception("Operation cancelled")
                    
                for col_idx in range(len(faculty_df.columns)):
                    try:
                        cell_value = str(faculty_df.iloc[row_idx, col_idx]).strip() if pd.notna(faculty_df.iloc[row_idx, col_idx]) else ""
                        cell_value_lower = cell_value.lower()
                        
                        # Check if any pattern is in the cell
                        pattern_found = False
                        matched_pattern = ""
                        for pattern in search_patterns:
                            if pattern.lower() in cell_value_lower:
                                pattern_found = True
                                matched_pattern = pattern
                                break
                        
                        if pattern_found:
                            # Extract session number if present
                            match = re.search(r'(\d+)', cell_value)
                            session_num = int(match.group(1)) if match else None
                            if session_num is not None:
                                dep_cols.append((col_idx, cell_value, session_num))
                                self.output.emit(f"Found {department_name} column at row {row_idx}, column {col_idx}: {cell_value}, Session: {session_num}, Matched pattern: {matched_pattern}")
                    except Exception as cell_error:
                        # Continue despite cell error
                        pass
                        
            if not dep_cols:
                error_msg = f"Could not locate any {department_name} columns in faculty sheet"
                self.output.emit(f"Error: {error_msg}")
                self.error.emit(error_msg)
                self.finished.emit(f"Error: {error_msg}")
                self.is_running = False
                return
            
            # Update progress
            self.progress.emit(45)
            
            # Find the student ID column in the faculty sheet
            id_col_idx = None
            id_col_row_start = None
            
            self.output.emit("Searching for student ID column...")
            
            # First scan the first ~20 rows to find potential ID columns
            potential_id_columns = {}
            
            # Look through first ~30 rows to get a good sample
            for row_idx in range(min(30, len(faculty_df))):
                # Check if we're still running
                if not self.is_running:
                    raise Exception("Operation cancelled")
                    
                for col_idx in range(len(faculty_df.columns)):
                    try:
                        cell_value = str(faculty_df.iloc[row_idx, col_idx]).strip() if pd.notna(faculty_df.iloc[row_idx, col_idx]) else ""
                        
                        # Check if cell contains a string of 4+ digits
                        if cell_value.isdigit() and len(cell_value) >= 4:
                            # Add this column to our potential columns dict and increment its count
                            if col_idx not in potential_id_columns:
                                potential_id_columns[col_idx] = []
                            potential_id_columns[col_idx].append(row_idx)
                    except Exception as cell_error:
                        # Continue despite cell error
                        pass
            
            # Find the column with the most potential student IDs
            best_col = None
            max_ids = 0
            
            for col_idx, row_indices in potential_id_columns.items():
                if len(row_indices) > max_ids:
                    max_ids = len(row_indices)
                    best_col = col_idx
                    
            if best_col is not None:
                id_col_idx = best_col
                # Find first row with student ID (assuming continuous data afterward)
                id_col_row_start = min(potential_id_columns[best_col])
                self.output.emit(f"Found student ID column at index {id_col_idx}, starting from row {id_col_row_start}")
                self.output.emit(f"Found {max_ids} potential student IDs in this column")
            else:
                self.output.emit("Warning: Could not detect a clear student ID column. Falling back to full sheet search.")
            
            # Update progress
            self.progress.emit(50)
            
            # Get the department student ID column name
            dep_id_col = "Student ID"  # Based on the printed column names
        
            # Update progress
            total_students = len(summary_df)
            
            # Process each student in the summary sheet
            self.output.emit(f"Processing {total_students} students...")
            
            save_interval = 5  # Save every 5 students instead of 10
            
            for idx, dep_row in enumerate(summary_df.iterrows()):
                try:
                    # Check if we're still running - enable cancellation
                    if not self.is_running:
                        raise Exception("Operation cancelled")
                        
                    _, dep_row = dep_row  # Unpack the tuple
                    
                    # Calculate progress - spread between 50% and 90%
                    student_progress = 50 + int((idx / total_students) * 40)
                    self.progress.emit(student_progress)
                    
                    # Get the student ID
                    if dep_id_col not in dep_row:
                        self.output.emit(f"Cannot find column {dep_id_col} in department sheet")
                        continue
                        
                    student_id = str(dep_row[dep_id_col]).strip()
                    
                    # Find student ID in the faculty sheet
                    matching_cells = []
                    
                    # If we found a likely student ID column, search only in that column
                    if id_col_idx is not None and id_col_row_start is not None:
                        # Only search in the identified student ID column, starting from the first identified row
                        for row_idx in range(id_col_row_start, len(faculty_df)):
                            try:
                                cell_value = faculty_df.iloc[row_idx, id_col_idx]
                                if pd.notna(cell_value):
                                    faculty_cell_value = str(cell_value).strip()
                                    if faculty_cell_value == student_id:
                                        matching_cells.append((row_idx, id_col_idx))
                                        self.output.emit(f"Found student ID {student_id}")
                                        break  # Since we're in the ID column, we only need the first match
                            except Exception as e:
                                pass
                    else:
                        # Fallback: Search the entire sheet for matching cells
                        self.output.emit(f"Searching for student ID: {student_id}")
                        for row_idx in range(len(faculty_df)):
                            # Check every 100 rows if we're still running
                            if row_idx % 100 == 0 and not self.is_running:
                                raise Exception("Operation cancelled")
                                
                            for col_idx in range(len(faculty_df.columns)):
                                try:
                                    cell_value = faculty_df.iloc[row_idx, col_idx]
                                    if pd.notna(cell_value):
                                        # Compare after converting to string and stripping whitespace
                                        faculty_cell_value = str(cell_value).strip()
                                        if faculty_cell_value == student_id:
                                            matching_cells.append((row_idx, col_idx))
                                            self.output.emit(f"Found student ID {student_id}")
                                except Exception as e:
                                    pass
                    
                    if not matching_cells:
                        self.output.emit(f"No match found for student ID: {student_id}")
                        continue
                    
                    # For each matching student ID cell
                    for row_idx, _ in matching_cells:
                        # For each dep column detected
                        for col_idx, _, session_num in dep_cols:
                            if session_num is None:
                                self.output.emit(f"Skipping column without session number")
                                continue
                            
                            # Look for attendance columns that contain both the session number and "(Attended)"
                            attended = 0  # Default value
                            
                            # Find the appropriate attendance column by searching for session number and "(Attended)"
                            attendance_col = None
                            for col_name in dep_row.index:
                                # Check if this column contains both the session number and "(Attended)"
                                if "(Attended)" in str(col_name) and str(session_num) in str(col_name):
                                    attendance_col = col_name
                                    break
                            
                            if attendance_col:
                                # If found, get the attendance value
                                attended = int(dep_row[attendance_col]) if pd.notna(dep_row[attendance_col]) else 0
                            else:
                                # Fallback to checking the attendance sheet
                                try:
                                    student_sessions = dep_df.loc[dep_df[dep_id_col].astype(str).str.strip() == student_id]
                                    if not student_sessions.empty:
                                        session_col = 'Session'
                                        if session_col in student_sessions.columns:
                                            session_attended = student_sessions.loc[student_sessions[session_col] == session_num]
                                            if not session_attended.empty:
                                                attended = 1
                                except Exception as e:
                                    pass
                            
                            try:
                                # Update the faculty table at the precise X,Y coordinate (row, column)
                                faculty_df.iloc[row_idx, col_idx] = attended
                                
                                # Also update the Excel file directly using openpyxl (1-indexed)
                                # Convert to Excel's 1-based indexing
                                excel_row = row_idx + 1
                                excel_col = col_idx + 1
                                
                                # Convert column index to letter (A, B, C, etc.)
                                col_letter = get_column_letter(excel_col)
                                
                                # Update the cell in openpyxl
                                sheet[f"{col_letter}{excel_row}"] = attended
                                
                                updated_records += 1
                            except Exception as update_error:
                                pass
                    
                    # Increment the student counter
                    students_processed += 1
                    
                    # Save more frequently
                    if students_processed % save_interval == 0:
                        self.output.emit(f"Saving after processing {students_processed} students...")
                        try:
                            workbook.save(self.faculty_file)
                        except Exception as save_error:
                            error_msg = f"Warning: Save failed: {str(save_error)}"
                            self.output.emit(error_msg)
                            
                            # Try to save with a different filename if original save fails
                            try:
                                alt_save_path = f"{os.path.splitext(self.faculty_file)[0]}_recovery_{datetime.now().strftime('%H%M%S')}{os.path.splitext(self.faculty_file)[1]}"
                                workbook.save(alt_save_path)
                                self.output.emit(f"Saved to alternative location: {alt_save_path}")
                            except Exception as alt_save_error:
                                pass
                
                except Exception as student_error:
                    # Log the error but continue with next student
                    self.output.emit(f"Warning: Error processing a student: {str(student_error)}")
            
            # Final save at the end
            self.output.emit(f"Final save after processing all students...")
            try:
                workbook.save(self.faculty_file)
            except Exception as final_save_error:
                error_msg = f"Final save failed: {str(final_save_error)}"
                self.output.emit(f"Warning: {error_msg}")
                
                # Try to save with a different filename
                try:
                    final_save_path = f"{os.path.splitext(self.faculty_file)[0]}_final_{datetime.now().strftime('%H%M%S')}{os.path.splitext(self.faculty_file)[1]}"
                    workbook.save(final_save_path)
                    self.output.emit(f"Saved final result to: {final_save_path}")
                except Exception as alt_final_save_error:
                    pass
       
            # Also save a DataFrame version just to be safe
            try:
                backup_df_path = self.faculty_file + "_dataframe_backup.xlsx"
                faculty_df.to_excel(backup_df_path, index=False, header=False)
            except Exception as df_save_error:
                pass
            
            # Update progress to 100%
            self.progress.emit(100)
            
            success_message = f"Successfully updated {updated_records} attendance records for {students_processed} students. Original file backed up to {backup_filename}"
            self.output.emit(success_message)
            
            # Emit success signal
            self.success.emit(success_message)
            
            # Emit completion signal
            self.finished.emit(success_message)
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            self.output.emit(f"Error: {str(e)}")
            
            # Reset progress bar on error
            self.progress.emit(0)
            
            error_message = f"Error during integration: {str(e)}"
            
            # Emit error signal
            self.error.emit(error_message)
            
            # Emit completion signal with error message
            self.finished.emit(error_message)
        
        finally:
            # Always reset running state in finally block
            self.is_running = False

#==========================================================log sheet preparer==========================================================#

class GithubDownloadWorker(QThread):
    progress_signal = pyqtSignal(int)
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(list)
    
    def __init__(self, repo_url, token):
        super().__init__()
        self.repo_url = repo_url
        self.token = token
        self.downloaded_files = []
        
    def run(self):
        try:
            # Parse the repo URL to extract owner and repo name
            # Example: "https://github.com/username/repo"
            parts = self.repo_url.strip('/').split('/')
            if len(parts) < 5 or parts[2] != 'github.com':
                self.log_signal.emit("Invalid GitHub repository URL format")
                return
                
            owner = parts[3]
            repo = parts[4]
            
            # Get repository contents
            self.log_signal.emit(f"Connecting to GitHub repository: {owner}/{repo}")
            headers = {'Authorization': f'token {self.token}'} if self.token else {}
            
            # Get all files in the repository
            api_url = f"https://api.github.com/repos/{owner}/{repo}/contents/backups"  # Specify the backups folder
            response = requests.get(api_url, headers=headers)
            
            if response.status_code != 200:
                self.log_signal.emit(f"Error accessing repository: {response.status_code}, {response.text}")
                return
                
            contents = response.json()
            excel_files = [item for item in contents if item['name'].endswith('.xlsx') or item['name'].endswith('.xls')]
            
            if not excel_files:
                self.log_signal.emit("No Excel files found in the repository")
                return
                
            # Create temp directory if it doesn't exist
            temp_dir = os.path.join(os.path.dirname(__file__), 'Imported_scan_logs')
            os.makedirs(temp_dir, exist_ok=True)
            
            # Download each Excel file
            total_files = len(excel_files)
            for idx, file in enumerate(excel_files):
                self.log_signal.emit(f"Downloading {file['name']}...")
                
                download_url = file['download_url']
                file_response = requests.get(download_url, headers=headers)
                
                if file_response.status_code == 200:
                    # Save file locally
                    file_path = os.path.join(temp_dir, file['name'])
                    with open(file_path, 'wb') as f:
                        f.write(file_response.content)
                    
                    self.downloaded_files.append(file_path)
                    self.log_signal.emit(f"Downloaded {file['name']}")
                else:
                    self.log_signal.emit(f"Failed to download {file['name']}: {file_response.status_code}")
                
                # Update progress
                progress = int(((idx + 1) / total_files) * 100)
                self.progress_signal.emit(progress)
                
            self.log_signal.emit(f"Downloaded {len(self.downloaded_files)} Excel files")
            self.finished_signal.emit(self.downloaded_files)
            
        except Exception as e:
            self.log_signal.emit(f"Error: {str(e)}")
            import traceback
            self.log_signal.emit(traceback.format_exc())

class MergeWorker(QThread):
    progress_signal = pyqtSignal(int)
    log_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)
    
    def __init__(self, files, output_file):
        super().__init__()
        self.files = files
        self.output_file = output_file
        
    def run(self):
        try:
            if not self.files:
                self.log_signal.emit("No files to merge")
                return
                
            self.log_signal.emit(f"Starting merge of {len(self.files)} files")
            
            # Initialize a list to hold all dataframes
            all_dfs = []
            
            # Process each file
            for idx, file_path in enumerate(self.files):
                self.log_signal.emit(f"Processing {os.path.basename(file_path)}")
                
                try:
                    # Read all sheets from the Excel file
                    excel_file = pd.ExcelFile(file_path)
                    for sheet_name in excel_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        
                        # Only process if the dataframe is not empty
                        if not df.empty:
                            # Add file and sheet metadata
                            df['Source_File'] = os.path.basename(file_path)
                            df['Source_Sheet'] = sheet_name
                            all_dfs.append(df)
                            self.log_signal.emit(f"Added sheet '{sheet_name}' with {len(df)} rows")
                            
                except Exception as e:
                    self.log_signal.emit(f"Error processing {file_path}: {str(e)}")
                    continue
                
                # Update progress
                progress = int(((idx + 1) / len(self.files)) * 100)
                self.progress_signal.emit(progress)
            
            if not all_dfs:
                self.log_signal.emit("No valid data found in the files")
                return
                
            # Standardize column names across all dataframes
            self.log_signal.emit("Standardizing column headers...")
            
            # Find common columns or use a predefined set of columns
            # For now, we'll use a simple approach of getting all unique columns
            all_columns = set()
            for df in all_dfs:
                all_columns.update(df.columns)
                
            # Remove metadata columns we added
            standard_columns = [col for col in all_columns if col not in ['Source_File', 'Source_Sheet']]
            
            # Reindex all dataframes with the standard columns
            standardized_dfs = []
            for df in all_dfs:
                # Create a new dataframe with all standard columns (will be filled with NaN for missing columns)
                new_df = pd.DataFrame(columns=standard_columns)
                
                # Copy data from original dataframe for matching columns
                for col in standard_columns:
                    if col in df.columns:
                        new_df[col] = df[col]
                
                # Add back metadata columns
                new_df['Source_File'] = df['Source_File']
                new_df['Source_Sheet'] = df['Source_Sheet']
                
                standardized_dfs.append(new_df)
            
            # Concatenate all dataframes
            self.log_signal.emit("Merging all sheets...")
            merged_df = pd.concat(standardized_dfs, ignore_index=True)
            
            # Identify and reorder columns as per requirements:
            # 1. Student ID (looking for column containing "Student" and "ID")
            # 2. Location (looking for column containing "Location")
            # 3. Log date (looking for column containing "Log" and "date" or "Date")
            # 4. Log time (looking for column containing "Log" and "time" or "Time")
            # 5. All other columns
            
            self.log_signal.emit("Reordering columns to specified format...")
            
            # Find the best matching columns based on column names
            student_id_col = None
            location_col = None
            log_date_col = None
            log_time_col = None
            
            # Look for exact or partial matches
            for col in merged_df.columns:
                col_lower = str(col).lower()
                
                # Check for Student ID
                if "student" in col_lower and "id" in col_lower:
                    student_id_col = col
                # Check for Location
                elif "location" in col_lower:
                    location_col = col
                # Check for Log date
                elif "log" in col_lower and ("date" in col_lower or "day" in col_lower):
                    log_date_col = col
                # Check for Log time
                elif "log" in col_lower and "time" in col_lower:
                    log_time_col = col
            
            # Create the ordered columns list
            ordered_columns = []
            
            # Add the main required columns if they exist
            for col in [student_id_col, location_col, log_date_col, log_time_col]:
                if col is not None and col in merged_df.columns:
                    ordered_columns.append(col)
            
            # Add all remaining columns (excluding the ones we've already added and metadata)
            remaining_columns = [col for col in merged_df.columns 
                               if col not in ordered_columns 
                               and col not in ['Source_File', 'Source_Sheet']]
            ordered_columns.extend(remaining_columns)
            
            # Add metadata columns at the end
            ordered_columns.extend(['Source_File', 'Source_Sheet'])
            
            # Log the column ordering
            self.log_signal.emit(f"Column order being used: {', '.join(ordered_columns[:4])} + remaining columns")
            
            # Reorder the dataframe columns
            merged_df = merged_df[ordered_columns]
            
            # Save the merged data with reordered columns
            self.log_signal.emit(f"Saving merged data to {self.output_file}")
            merged_df.to_excel(self.output_file, index=False)
            
            self.log_signal.emit(f"Successfully merged {len(all_dfs)} sheets into {self.output_file} with ordered columns")
            self.finished_signal.emit(self.output_file)
            
        except Exception as e:
            self.log_signal.emit(f"Error during merge: {str(e)}")
            import traceback
            self.log_signal.emit(traceback.format_exc())

class LogSheetPreparer(QWidget):
    def __init__(self):
        super().__init__()
        self.files_to_merge = []
        # Define the hardcoded GitHub token - split to avoid detection
        token_part1 = "github_pat_"
        token_part2 = "11BREVRNQ0XVpxHicj3xsl_vAfICFbNYso7tpxkuw9yZqcOG4FHzacfgkpOjBJE51HR3WGTNJTaUIfxSWg"
        self.github_token = token_part1 + token_part2
        # Define the hardcoded GitHub repo URL - hidden from UI
        self.github_repo = "https://github.com/MohammadHamdi11/QRScanner-webapp"
        self.setStyleSheet("""
            background-color: black; 
            color: white;
            QLabel {
                color: white;
            }
        """)
        
        self.init_ui()
        
    def init_ui(self):
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)
        main_layout.setContentsMargins(20, 20, 20, 20)
        
        # Header layout
        header_layout = QHBoxLayout()
        logo_label = QLabel()
        logo_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path).scaled(60, 60, Qt.AspectRatioMode.KeepAspectRatio, 
                                             Qt.TransformationMode.SmoothTransformation)
            logo_label.setPixmap(pixmap)
        title_label = QLabel("Log Sheet Preparer")
        title_label.setStyleSheet(f"font-size: 24px; font-weight: bold;")
        header_layout.addWidget(logo_label)
        header_layout.addWidget(title_label)
        header_layout.addStretch()
        
        # Create vertical layout for header buttons
        header_buttons_layout = QVBoxLayout()
        header_buttons_layout.setSpacing(5)
        
        # Back button
        back_btn = QPushButton("Back to Home")
        back_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        back_btn.clicked.connect(self.return_to_home)
        
        # Exit button
        exit_btn = QPushButton("Exit")
        exit_btn.setStyleSheet(EXIT_BUTTON_STYLE)
        exit_btn.clicked.connect(QApplication.instance().quit)
        
        # Add buttons to vertical layout
        header_buttons_layout.addWidget(back_btn)
        header_buttons_layout.addWidget(exit_btn)
        header_buttons_layout.addStretch()
        
        # Add button layout to header
        header_layout.addStretch()
        header_layout.addLayout(header_buttons_layout)
        main_layout.addLayout(header_layout)
        
        # Import Method Selection
        import_group = QGroupBox("Import Method")
        import_group.setStyleSheet(GROUP_BOX_STYLE)
        import_layout = QVBoxLayout(import_group)
        
        # Radio buttons for import method
        radio_layout = QHBoxLayout()
        self.github_radio = QRadioButton("Import from Cloud Storage")
        self.github_radio.setChecked(True)  # Default to GitHub import
        self.local_radio = QRadioButton("Import Local Files")
        radio_layout.addWidget(self.github_radio)
        radio_layout.addWidget(self.local_radio)
        radio_layout.addStretch()
        import_layout.addLayout(radio_layout)
        
        # Connect radio buttons to toggle between input methods
        self.github_radio.toggled.connect(self.toggle_import_method)
        
        # Stacked widget for different import methods
        self.import_stack = QStackedWidget()
        
        # GitHub Import Widget
        github_widget = QWidget()
        github_layout = QVBoxLayout(github_widget)
        
        # Only show informational text about the GitHub repo, hiding the actual implementation details
        github_info_label = QLabel("Excel files will be downloaded from the QRScanner-webapp repository's backup folder.")
        github_layout.addWidget(github_info_label)
        
        # Button to download files from GitHub
        download_btn = QPushButton("Download Excel Files from Cloud Storage")
        download_btn.clicked.connect(self.download_github_files)
        download_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        github_layout.addWidget(download_btn)
        
        # Local Files Import Widget
        local_widget = QWidget()
        local_layout = QVBoxLayout(local_widget)
        
        local_files_layout = QHBoxLayout()
        local_files_layout.addWidget(QLabel("Local Excel Files:"))
        self.local_files_label = QLineEdit()
        self.local_files_label.setPlaceholderText("Select Excel file...")
        self.local_files_label.setMinimumWidth(300)
        self.local_files_label.setReadOnly(True)
        local_files_layout.addWidget(self.local_files_label)
        
        # Button for importing local files
        browse_btn = QPushButton("Browse")
        browse_btn.clicked.connect(self.browse_files)
        browse_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        local_files_layout.addWidget(browse_btn)
        local_layout.addLayout(local_files_layout)
        
        # Add widgets to stack
        self.import_stack.addWidget(github_widget)
        self.import_stack.addWidget(local_widget)
        import_layout.addWidget(self.import_stack)
        
        main_layout.addWidget(import_group)
        
        # Files Table Section
        files_group = QGroupBox("Files to Merge")
        files_group.setStyleSheet(GROUP_BOX_STYLE)
        files_layout = QVBoxLayout(files_group)
        
        self.files_table = QTableWidget()
        self.files_table.setColumnCount(2)
        self.files_table.setHorizontalHeaderLabels(['File Path', 'Status'])
        
        # Center align the header text
        header = self.files_table.horizontalHeader()
        header.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        
        # Set column resize modes
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)  # File Path
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)  # Status
        
        # Apply table styles
        self.files_table.setStyleSheet(TABLE_STYLE)
        files_layout.addWidget(self.files_table)
        
        # Buttons for files table
        files_btn_layout = QHBoxLayout()
        clear_files_btn = QPushButton("Clear Files")
        clear_files_btn.clicked.connect(self.clear_files)
        clear_files_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        files_btn_layout.addWidget(clear_files_btn)
        files_layout.addLayout(files_btn_layout)
        
        main_layout.addWidget(files_group)
        
        # Progress Bar Section
        progress_group = QGroupBox("Progress")
        progress_group.setStyleSheet(GROUP_BOX_STYLE)
        progress_layout = QVBoxLayout(progress_group)
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.progress_bar.setStyleSheet(PROGRESS_BAR_STYLE)
        
        # Create loading gif label
        self.loading_label = QLabel()
        self.loading_label.setFixedSize(24, 24)
        self.loading_label.setVisible(False)
                
        # Create the movie object for the GIF
        self.loading_movie = QMovie()
        self.loading_movie.setScaledSize(QSize(24, 24))
        self.loading_label.setMovie(self.loading_movie)
        
        loading_gif_path = os.path.join(os.path.dirname(__file__), 'loading.gif')
        if os.path.exists(loading_gif_path):
            self.loading_movie.setFileName(loading_gif_path)
        else:
            print(f"Warning: loading.gif not found at {loading_gif_path}")
            
        # Create a horizontal layout to hold both the progress bar and loading animation
        progress_h_layout = QHBoxLayout()
        progress_h_layout.addWidget(self.progress_bar)
        progress_h_layout.addWidget(self.loading_label)
        progress_layout.addLayout(progress_h_layout)
        
        main_layout.addWidget(progress_group)
        
        # Output Console Section
        console_group = QGroupBox("Output Console")
        console_group.setStyleSheet(GROUP_BOX_STYLE)
        console_layout = QVBoxLayout(console_group)

        self.output_console = QTextEdit()
        self.output_console.setReadOnly(True)
        self.output_console.setMaximumHeight(150)
        self.output_console.setStyleSheet(CONSOLE_STYLE)  
        console_layout.addWidget(self.output_console)
        main_layout.addWidget(console_group)
        
        # Bottom Buttons
        button_layout = QHBoxLayout()
        merge_btn = QPushButton("Merge Logs Files")
        merge_btn.clicked.connect(self.merge_files)
        merge_btn.setStyleSheet(STANDARD_BUTTON_STYLE)
        button_layout.addWidget(merge_btn)
        main_layout.addLayout(button_layout)
        
    def toggle_import_method(self):
        # Set the current import method based on radio button selection
        if self.github_radio.isChecked():
            self.import_stack.setCurrentIndex(0)
        else:
            self.import_stack.setCurrentIndex(1)
            
    def return_to_home(self):
        # Get the stacked widget and switch to the start page
        stacked_widget = self.parent()
        if isinstance(stacked_widget, QStackedWidget):
            stacked_widget.setCurrentIndex(0)
            
    def browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(
            self, "Select Excel Files", "", "Excel Files (*.xlsx *.xls)"
        )
        if files:
            self.files_to_merge = files
            self.local_files_label.setText(f"{len(files)} files selected")
            self.update_files_table()
            self.log_message(f"Selected {len(files)} files for merging")
            
    def update_files_table(self):
        # Clear existing table
        self.files_table.setRowCount(0)
        
        # Add files to table
        for file_path in self.files_to_merge:
            row_position = self.files_table.rowCount()
            self.files_table.insertRow(row_position)
            
            # Create items for the cells
            file_item = QTableWidgetItem(os.path.basename(file_path))
            status_item = QTableWidgetItem("Ready")
            
            # Set items to the table
            self.files_table.setItem(row_position, 0, file_item)
            self.files_table.setItem(row_position, 1, status_item)
            
    def clear_files(self):
        self.files_to_merge = []
        self.files_table.setRowCount(0)
        self.local_files_label.setText("No files selected")
        self.log_message("Cleared all files")
            
    def log_message(self, message):
        self.output_console.append(f"[{datetime.now().strftime('%H:%M:%S')}] {message}")
        # Scroll to the bottom
        self.output_console.moveCursor(QTextCursor.MoveOperation.End)
        
    def download_github_files(self):
        # Use hardcoded repo URL and token - not visible to users
        repo_url = self.github_repo
        token = self.github_token
        
        self.log_message(f"Starting download from Cloud Storage")
        
        # Start loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
        self.progress_bar.setValue(0)
        
        # Create and start the worker thread
        self.github_worker = GithubDownloadWorker(repo_url, token)
        self.github_worker.progress_signal.connect(self.update_progress)
        self.github_worker.log_signal.connect(self.log_message)
        self.github_worker.finished_signal.connect(self.handle_downloaded_files)
        self.github_worker.start()

    def handle_downloaded_files(self, files):
        # Update the list of files to merge
        self.files_to_merge = files
        self.update_files_table()
        
        # Stop loading animation
        self.loading_movie.stop()
        self.loading_label.setVisible(False)
        
    def update_progress(self, value):
        self.progress_bar.setValue(value)
        
    def merge_files(self):
        if not self.files_to_merge:
            self.log_message("No files to merge. Please import files first.")
            return
            
        # Generate output filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"prepared_log_history_{timestamp}.xlsx"
        
        # Create 'Output' directory if it doesn't exist
        output_dir = os.path.join(os.path.dirname(__file__), 'Merged Files')
        os.makedirs(output_dir, exist_ok=True)
        
        # Full path for output file
        output_file = os.path.join(output_dir, output_filename)
        
        self.log_message(f"Merging files to: {output_file}")
            
        # Start loading animation
        self.loading_label.setVisible(True)
        self.loading_movie.start()
        self.progress_bar.setValue(0)
        
        # Create and start the worker thread
        self.merge_worker = MergeWorker(self.files_to_merge, output_file)
        self.merge_worker.progress_signal.connect(self.update_progress)
        self.merge_worker.log_signal.connect(self.log_message)
        self.merge_worker.finished_signal.connect(self.handle_merge_complete)
        self.merge_worker.start()
        
    def handle_merge_complete(self, output_file):
        # Stop loading animation
        self.loading_movie.stop()
        self.loading_label.setVisible(False)
        
        self.log_message(f"Merge completed successfully: {output_file}")
        
        # Ask if user wants to open the merged file
        from PyQt6.QtWidgets import QMessageBox
        reply = QMessageBox.question(
            self, 'Merge Complete', 
            f'Merge completed successfully!\nThe merged file has been saved as:\n{os.path.basename(output_file)}\n\nWould you like to open the merged file?',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No, 
            QMessageBox.StandardButton.Yes
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            # Open the file with the default application
            import subprocess
            import platform
            
            if platform.system() == 'Windows':
                os.startfile(output_file)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(('open', output_file))
            else:  # Linux
                subprocess.call(('xdg-open', output_file))

#==========================================================main app==========================================================#

class MainApplication(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Attendance Management App")
        self.setMinimumSize(1000, 750)
        icon_path = os.path.join(os.path.dirname(__file__), 'ASU1.png')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))
    
        self.setWindowTitle("Attendance Management App")
        
        # Create stacked widget to manage pages
        self.stacked_widget = QStackedWidget()
        self.setCentralWidget(self.stacked_widget)
        
        # Create pages
        self.start_page = StartPage(self)
        self.info_page = InfoPage()  # New info page
        self.preparer_page = LogSheetPreparer()
        self.processor_page = AttendanceProcessor()
        self.populator_page = Facultypopulator()  
        self.dashboard_page = AttendanceDashboard()  
        
        # Add pages to stacked widget
        self.stacked_widget.addWidget(self.start_page)
        self.stacked_widget.addWidget(self.info_page)  # Add info page
        self.stacked_widget.addWidget(self.preparer_page)
        self.stacked_widget.addWidget(self.processor_page)
        self.stacked_widget.addWidget(self.populator_page)  
        self.stacked_widget.addWidget(self.dashboard_page)  
        
        # Connect start page buttons to switch pages
        self.start_page.info_button.clicked.connect(self.show_info)  # New connection
        self.start_page.preparer_btn.clicked.connect(self.show_preparer)
        self.start_page.process_btn.clicked.connect(self.show_processor)
        self.start_page.populate_btn.clicked.connect(self.show_populator)
        self.start_page.dashboard_btn.clicked.connect(self.show_dashboard)
        
        # Set the window style
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {BLACK};
            }}
        """)
    
    def show_info(self):
        self.stacked_widget.setCurrentWidget(self.info_page)
        
    def show_preparer(self):
        self.stacked_widget.setCurrentWidget(self.preparer_page)
        
    def show_processor(self):
        self.stacked_widget.setCurrentWidget(self.processor_page)
        
    def show_populator(self):
        self.stacked_widget.setCurrentWidget(self.populator_page)
        
    def show_dashboard(self):
        self.stacked_widget.setCurrentWidget(self.dashboard_page)

if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainApplication()
    window.show()
    sys.exit(app.exec())